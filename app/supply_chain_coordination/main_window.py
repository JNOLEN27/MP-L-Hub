import json
import re
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTabWidget, QTextEdit, QTableWidget, QTableWidgetItem, QMessageBox, QScrollArea, QFileDialog, QComboBox, QListWidget, QListWidgetItem, QCheckBox, QFrame, QApplication, QLineEdit, QGridLayout, QProgressDialog, QTableView, QMenu, QAction, QStyledItemDelegate)
from PyQt5.QtCore import Qt, pyqtSignal, QTimer, QEvent
from PyQt5.QtGui import QFont, QColor, QBrush
from datetime import datetime, timedelta
from app.utils.config import APPWINDOWSIZE, getsharednetworkpath, ADMINUSERS, LOCALAPPDATA
from app.data.import_manager import DataImportManager
from app.supply_chain_coordination.coverage_analysis import CoverageAnalysisEngine
from app.supply_chain_coordination.waterfall_analysis import WaterfallAnalysisEngine
from app.supply_chain_coordination.ldjis_coverage import LDJISCoverageEngine
 
class PersistentMenu(QMenu):
    def mouseReleaseEvent(self, event):
        action = self.activeAction()
        if action and action.isCheckable():
            action.trigger()
            return
        super().mouseReleaseEvent(event)

class NumericSortTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, sort_value):
        super().__init__(text)
        self.sort_value = sort_value

    def __lt__(self, other):
        if isinstance(other, NumericSortTableWidgetItem):
            return self.sort_value < getattr(other, "sort_value", 0)
        return super().__lt__(other)
 
class _BackgroundDelegate(QStyledItemDelegate):
    """Forces QTableWidgetItem BackgroundRole to render even when a stylesheet is applied."""
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        brush = index.data(Qt.BackgroundRole)
        if brush is not None and isinstance(brush, QBrush) and brush.style() != Qt.NoBrush:
            option.backgroundBrush = brush


class SupplyChainCoordinationWindow(QMainWindow):
    def __init__(self, userdata, parent=None):
        super().__init__(parent)
        self.userdata = userdata
        self.import_manager = DataImportManager()
        self.coverageengine = CoverageAnalysisEngine(self.import_manager)
        self.waterfallengine = WaterfallAnalysisEngine(self.import_manager)
        self.ldjiscoverageengine = LDJISCoverageEngine(self.import_manager)
        self.ldjismodelnames = ['EX90', 'PS3']
        self._comments_cache = None
        self._coverage_pending = None
        self._alerts_cache = None
        self._alerts_pending = None
        self._piwd_cache = None
        self._alert_highlights = None
        self.setWindowTitle("Supply Chain Coordination Application")
        self._dropdowns = []
        self._hidden_coverage_columns = set()
        self._coverage_column_menu = None
        self._current_screen = QApplication.primaryScreen()
        self._recalc_filter_heights(self._current_screen)
        self.resize(*self._screen_fitted_size(self._current_screen))
        self.setupui()
 
    def setupui(self):
        centralwidget = QWidget()
        self.setCentralWidget(centralwidget)
 
        centralwidget.setStyleSheet("QWidget#centralwidget { background-color: #156082; }")
        centralwidget.setObjectName("centralwidget")
 
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
 
        headerwidget = QWidget()
        headerwidget.setStyleSheet("background-color: #156082;")
        headerlayout = QVBoxLayout(headerwidget)
        headerlayout.setContentsMargins(10, 8, 10, 8)
 
        title = QLabel("Supply Chain Coordination")
        title.setFont(QFont("Arial", max(11, self._sz(18)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: white; background-color: transparent;")
        headerlayout.addWidget(title)

        layout.addWidget(headerwidget)

        self.setMinimumWidth(max(700, self._sz(900)))

        tabs = QTabWidget()
        tabs.tabBar().setElideMode(Qt.ElideNone)
        tabs.tabBar().setExpanding(False)
        self._tabs = tabs
        self._apply_tab_stylesheet()
 
        coveragedashtab = self.createcoveragedashboard()
        coverageindivtab = self.createindividualpartcoverage()
        calloffforecasttab = self.createcalloffforecasttab()
        ldjiscoveragetab = self.createldjiscoveragetab()
        alertstab = self.createalertstab()

        tabs.addTab(coveragedashtab, "Coverage Dashboard")
        tabs.addTab(coverageindivtab, "Individual Part Coverage")
        tabs.addTab(calloffforecasttab, "Call-off Forecast and Waterfall")
        tabs.addTab(ldjiscoveragetab, "LDJIS Coverage")
        tabs.addTab(alertstab, "Alerts Breakdown")

        from app.supply_chain_coordination.maintenance_tab import MaintenanceTab
        mainttab = MaintenanceTab(self.import_manager, self.userdata)
        tabs.addTab(mainttab, "Maintenance")

        layout.addWidget(tabs)
 
        self.statusBar().showMessage(f"Logged in as: {self.userdata['username']}")
        centralwidget.setLayout(layout)

    def _screen_fitted_size(self, screen):
        if screen:
            avail = screen.availableGeometry()
            w = max(900, min(APPWINDOWSIZE[0], int(avail.width() * 0.90)))
            h = max(650, min(APPWINDOWSIZE[1], int(avail.height() * 0.90)))
            return (w, h)
        return APPWINDOWSIZE

    def _recalc_filter_heights(self, screen):
        # Scale relative to the 1920×1200 @ 100% reference layout that defines
        # the intended proportions.  On a smaller or higher-DPI screen the
        # filter section stays the same fraction of the window as on the
        # reference display — not the same absolute pixel count.
        if screen:
            avail = screen.availableGeometry()
            scale = min(avail.width() / 1920.0, avail.height() / 1080.0)
            scale = max(0.60, min(1.0, scale))
        else:
            scale = 1.0
        self._ui_scale = scale
        # Dropdown height independent of filter section — just fits the list widget.
        self._dropdown_h = max(70, int(145 * scale))
        # Coverage Dashboard filter needs room for 2 rows of scaled search filters.
        search_row_h = max(45, int(85 * scale))
        self._filter_section_h = max(self._dropdown_h + 20, 2 * search_row_h + 15)
        # Alerts/PIWD filter sections only contain dropdowns — use tighter height.
        self._alert_filter_h = self._dropdown_h + 20

    def _sz(self, n):
        return max(1, int(n * self._ui_scale))

    def showEvent(self, event):
        super().showEvent(event)
        if not getattr(self, '_screen_signal_connected', False):
            win = self.windowHandle()
            if win is not None:
                win.screenChanged.connect(self._on_screen_changed)
                self._screen_signal_connected = True
        screen = QApplication.screenAt(self.geometry().center()) or QApplication.primaryScreen()
        if screen and screen is not getattr(self, '_current_screen', None):
            self._connect_screen_dpi(screen)
        if screen:
            self._recalc_filter_heights(screen)
            self._reapply_filter_heights()

    def _connect_screen_dpi(self, screen):
        old = getattr(self, '_current_screen', None)
        if old and old is not screen:
            try:
                old.logicalDotsPerInchChanged.disconnect(self._on_dpi_changed)
            except Exception:
                pass
        self._current_screen = screen
        try:
            screen.logicalDotsPerInchChanged.connect(self._on_dpi_changed)
        except Exception:
            pass

    def _on_screen_changed(self, new_screen):
        self._connect_screen_dpi(new_screen)
        self._recalc_filter_heights(new_screen)
        self._reapply_filter_heights()
        # Shrink window if it overflows the new display's available area.
        avail = new_screen.availableGeometry()
        sz = self.size()
        new_w = min(sz.width(), int(avail.width() * 0.95))
        new_h = min(sz.height(), int(avail.height() * 0.95))
        if new_w != sz.width() or new_h != sz.height():
            self.resize(new_w, new_h)

    def _on_dpi_changed(self, _dpi):
        screen = getattr(self, '_current_screen', None) or QApplication.primaryScreen()
        self._recalc_filter_heights(screen)
        self._reapply_filter_heights()
        fitted = self._screen_fitted_size(screen)
        sz = self.size()
        if sz.width() > fitted[0] or sz.height() > fitted[1]:
            self.resize(min(sz.width(), fitted[0]), min(sz.height(), fitted[1]))

    def _apply_tab_stylesheet(self):
        if not hasattr(self, '_tabs'):
            return
        s = self._ui_scale
        tab_font = max(8, self._sz(11))
        tab_pad_v = max(3, self._sz(6))
        tab_pad_h = max(5, self._sz(14))
        tab_min_w = max(70, self._sz(150))
        self._tabs.setStyleSheet(f"""
            QTabWidget {{ background-color: #CCECFF; }}
            QTabWidget::pane {{
                background-color: white;
                border: 1px solid #99CCEE;
                border-top: 0px;
            }}
            QTabBar {{ background-color: #156082; }}
            QTabBar QToolButton {{
                background-color: #156082;
                border: 1px solid #99CCEE;
                color: white;
            }}
            QTabBar::tab {{
                background-color: white;
                color: #1A3A6B;
                font-weight: bold;
                font-size: {tab_font}px;
                padding: {tab_pad_v}px {tab_pad_h}px;
                min-width: {tab_min_w}px;
                border: 1px solid #99CCEE;
                border-bottom: none;
                border-radius: 4px 4px 0px 0px;
                margin-right: 2px;
            }}
            QTabBar::tab:selected {{
                background-color: #CCECFF;
                color: black;
                border: 1px solid #99CCEE;
                border-bottom: 1px solid #CCECFF;
            }}
            QTabBar::tab:hover:!selected {{ background-color: #E8F6FF; }}
        """)

    def _reapply_filter_heights(self):
        if hasattr(self, 'filtersection'):
            self.filtersection.setMaximumHeight(self._filter_section_h)
        if hasattr(self, 'alertfiltersection'):
            self.alertfiltersection.setMaximumHeight(self._alert_filter_h)
        if hasattr(self, 'piwdfiltersection'):
            self.piwdfiltersection.setMaximumHeight(self._alert_filter_h)
        for dd in self._dropdowns:
            dd.setFixedHeight(self._dropdown_h)
        self._apply_tab_stylesheet()

    def createfiltersection(self):
        widget = QWidget()
        widget.setMaximumHeight(self._filter_section_h)

        layout = QHBoxLayout()

        filterlabel = QLabel("Filters:")
        filterlabel.setFont(QFont("Arial", max(8, self._sz(12)), QFont.Bold))
        filterlabel.setAlignment(Qt.AlignTop)
        layout.addWidget(filterlabel)
 
        self.sccfilter = self.createmultiselectdropdown("Select SCC Names...")
        self.sccfilter.selectionChanged.connect(self.applyfilters)
        layout.addWidget(self.sccfilter)
 
        self.regionfilter = self.createmultiselectdropdown("Select Regions...", "Region")
        self.regionfilter.selectionChanged.connect(self.applyfilters)
        layout.addWidget(self.regionfilter)
 
        self.dayalertfilter = self.createmultiselectdropdown("Select Day Alert...", "Day Alert")
        self.dayalertfilter.selectionChanged.connect(self.applyfilters)
        layout.addWidget(self.dayalertfilter)

        self.programsupportedfilter = self.createmultiselectdropdown("Select Program...", "Program")
        self.programsupportedfilter.selectionChanged.connect(self.applyfilters)
        layout.addWidget(self.programsupportedfilter)
 
        searchgridwidget = QWidget()
        searchgridlayout = QGridLayout(searchgridwidget)
        searchgridlayout.setSpacing(5)
 
        self.searchfilters = []
 
        partfilter = self.createsearchfilter("Part", "PART_NO")
        self.searchfilters.append(partfilter)
        searchgridlayout.addWidget(partfilter, 0, 0)
 
        mfgfilter = self.createsearchfilter("MFG", "SUPP_MFG")
        self.searchfilters.append(mfgfilter)
        searchgridlayout.addWidget(mfgfilter, 0, 1)
 
        shpfilter = self.createsearchfilter("SHP", "SUPP_SHP")
        self.searchfilters.append(shpfilter)
        searchgridlayout.addWidget(shpfilter, 1, 0)
 
        countryfilter = self.createsearchfilter("Country", "SUPP_SHP_COUNTRY")
        self.searchfilters.append(countryfilter)
        searchgridlayout.addWidget(countryfilter, 1, 1)
 
        layout.addWidget(searchgridwidget)

        partlistwidget = QWidget()
        partlistlayout = QVBoxLayout(partlistwidget)
        partlistlayout.setContentsMargins(5, 5, 5, 5)
        partlistlayout.setSpacing(3)

        partlistheader = QHBoxLayout()
        partlistlabel = QLabel("Multiple Part Search:")
        partlistlabel.setFont(QFont("Arial", max(7, self._sz(10)), QFont.Bold))
        partlistheader.addWidget(partlistlabel)
        partlistheader.addStretch()
        clearpartlistbtn = QPushButton("✕")
        clearpartlistbtn.setFixedSize(max(15, self._sz(20)), max(15, self._sz(20)))
        clearpartlistbtn.setToolTip("Clear part list filter")
        clearpartlistbtn.clicked.connect(self._clearpartlistfilter)
        partlistheader.addWidget(clearpartlistbtn)
        partlistlayout.addLayout(partlistheader)

        self._partlist_textedit = QTextEdit()
        self._partlist_textedit.setPlaceholderText(
            "Paste part numbers here\n(one per line or comma-separated)"
        )
        self._partlist_textedit.setFixedHeight(max(60, self._sz(115)))
        self._partlist_textedit.setFixedWidth(max(120, self._sz(175)))
        self._partlist_textedit.setStyleSheet(f"font-size: {max(7, self._sz(10))}px; font-family: monospace;")
        partlistlayout.addWidget(self._partlist_textedit)

        self._partlist_statuslabel = QLabel("")
        self._partlist_statuslabel.setStyleSheet(f"font-size: {max(7, self._sz(9))}px; color: #555;")
        partlistlayout.addWidget(self._partlist_statuslabel)

        layout.addWidget(partlistwidget)

        self._partlist_timer = QTimer()
        self._partlist_timer.setSingleShot(True)
        self._partlist_timer.setInterval(400)
        self._partlist_timer.timeout.connect(self.applyfilters)
        self._partlist_textedit.textChanged.connect(self._partlist_timer.start)

        clearfiltersbtn = QPushButton("Clear All Filters")
        clearfiltersbtn.clicked.connect(self.clearfilters)
        clearfiltersbtn.setMaximumWidth(120)
        clearfiltersbtn.setMaximumHeight(30)
        layout.addWidget(clearfiltersbtn)
 
        layout.addStretch()
        widget.setLayout(layout)
        return widget
 
    def createalertfiltersection(self):
        widget = QWidget()
        widget.setMaximumHeight(self._alert_filter_h)
 
        layout = QHBoxLayout()
 
        filterlabel = QLabel("Filters:")
        filterlabel.setFont(QFont("Arial", max(8, self._sz(12)), QFont.Bold))
        filterlabel.setAlignment(Qt.AlignTop)
        layout.addWidget(filterlabel)
 
        self.alerts_scc_filter = self.createmultiselectdropdown("Select SCC Names...", "SCC")
        self.alerts_scc_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_scc_filter)
 
        self.alerts_type_filter = self.createmultiselectdropdown("Select Alert...", "Alert")
        self.alerts_type_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_type_filter)
 
        self.alerts_part_filter = self.createmultiselectdropdown("Select Part-Alert...", "Part-Alert")
        self.alerts_part_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_part_filter)

        self.alerts_region_filter = self.createmultiselectdropdown("Select Region...", "Region")
        self.alerts_region_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_region_filter)

        self.alerts_country_filter = self.createmultiselectdropdown("Select Country...", "Country")
        self.alerts_country_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_country_filter)

        self.alerts_program_filter = self.createmultiselectdropdown("Select Program...", "Program")
        self.alerts_program_filter.selectionChanged.connect(self.applyalertfilters)
        layout.addWidget(self.alerts_program_filter)

        clearfiltersbtn = QPushButton("Clear All Filters")
        clearfiltersbtn.clicked.connect(self.clearalertfilters)
        clearfiltersbtn.setMaximumWidth(120)
        clearfiltersbtn.setMaximumHeight(30)
        layout.addWidget(clearfiltersbtn)
 
        layout.addStretch()
        widget.setLayout(layout)
        return widget
 
    def createmultiselectdropdown(self, placeholdertext, filtertype="SCC"):
        dropdown_h = self._dropdown_h
        label_font_pt = max(7, self._sz(10))
        list_font_px = max(8, self._sz(11))
        select_font_pt = max(7, self._sz(9))
        class SimpleMultiSelectFilter(QWidget):
            selectionChanged = pyqtSignal()

            def __init__(self, placeholder="Select items...", filtertype="SCC"):
                super().__init__()
                self.filtertype = filtertype
                self.setFixedHeight(dropdown_h)

                layout = QVBoxLayout()
                layout.setContentsMargins(5, 5, 5, 5)

                self.label = QLabel(f"{filtertype} Filter:")
                self.label.setFont(QFont("Arial", label_font_pt, QFont.Bold))
                layout.addWidget(self.label)

                self.list_widget = QListWidget()
                self.list_widget.itemChanged.connect(self.on_item_changed)
                self.list_widget.setStyleSheet(f"""
                    QListWidget {{
                        border: 1px solid #ccc;
                        font-size: {list_font_px}px;
                    }}
                    QListWidget::item {{
                        padding: 2px;
                    }}
                """)
                layout.addWidget(self.list_widget)

                self.setLayout(layout)
                self.selected_items = set()
                self.all_items = []

            def additems(self, items, presorted=False):
                self.list_widget.clear()
                self.selected_items.clear()
                self.all_items.clear()

                select_all_item = QListWidgetItem(f"✓ All {self.filtertype}'s")
                select_all_item.setFlags(select_all_item.flags() | Qt.ItemIsUserCheckable)
                select_all_item.setCheckState(Qt.Checked)
                select_all_item.setData(Qt.UserRole, "SELECT_ALL")
                select_all_item.setFont(QFont("Arial", select_font_pt, QFont.Bold))
                self.list_widget.addItem(select_all_item)
 
                maxtextlength = len("✓ All SCC Names")
 
                items_to_use = items if presorted else sorted(items)
 
                for item in items_to_use:
                    if item and str(item).strip():
                        list_item = QListWidgetItem(str(item))
                        list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
                        list_item.setCheckState(Qt.Checked)
                        self.list_widget.addItem(list_item)
                        self.selected_items.add(str(item))
                        self.all_items.append(str(item))
                        maxtextlength = max(maxtextlength, len(str(item)))
 
                self.autosizewidth(maxtextlength)
                self.update_label()
 
            def autosizewidth(self, maxtextlength):
                fontmetrics = self.list_widget.fontMetrics()
                estimatedwidth = fontmetrics.averageCharWidth() * maxtextlength + 60
 
                if self.parent():
                    parentwidth = self.parent().width()
                    availablewidth = parentwidth - 280
                else:
                    availablewidth = 400
 
                minwidth = 150
                maxwidth = min(400, availablewidth)
                optimalwidth = max(minwidth, min(estimatedwidth, maxwidth))
 
                self.setFixedWidth(optimalwidth)
 
            def on_item_changed(self, item):
                self.list_widget.blockSignals(True)
 
                try:
                    if item.data(Qt.UserRole) == "SELECT_ALL":
                        if item.checkState() == Qt.Checked:
                            for i in range(1, self.list_widget.count()):
                                other_item = self.list_widget.item(i)
                                other_item.setCheckState(Qt.Checked)
                                self.selected_items.add(other_item.text())
                        else:
                            for i in range(1, self.list_widget.count()):
                                other_item = self.list_widget.item(i)
                                other_item.setCheckState(Qt.Unchecked)
                            self.selected_items.clear()
                    else:
                        if item.checkState() == Qt.Checked:
                            self.selected_items.add(item.text())
                        else:
                            self.selected_items.discard(item.text())
                            select_all_item = self.list_widget.item(0)
                            select_all_item.setCheckState(Qt.Unchecked)
 
                        if len(self.selected_items) == len(self.all_items):
                            select_all_item = self.list_widget.item(0)
                            select_all_item.setCheckState(Qt.Checked)
 
                finally:
                    self.list_widget.blockSignals(False)
 
                self.update_label()
                self.selectionChanged.emit()
 
            def show_progress_for_large_operation(self):
                progress = QProgressDialog("Loading full dataset...", "Cancel", 0, 100, self)
                progress.setWindowTitle("Processing")
                progress.setWindowModality(Qt.WindowModal)
                progress.setValue(50)
                progress.show()
                QApplication.processEvents()
                return progress
 
            def update_label(self):
                self.label.setText(f"{self.filtertype} Filter:")
 
            def getselecteditems(self):
                return list(self.selected_items)
 
            def selectallitems(self):
                if len(self.all_items) > 50:
                    progress = self.show_progress_for_large_operation()
                else:
                    progress = None
 
                try:
                    self.list_widget.blockSignals(True)
                    self.selected_items = set(self.all_items)
                    for i in range(self.list_widget.count()):
                        item = self.list_widget.item(i)
                        item.setCheckState(Qt.Checked)
                finally:
                    self.list_widget.blockSignals(False)
                    if progress:
                        progress.close()
 
                self.update_label()
 
        widget = SimpleMultiSelectFilter(placeholdertext, filtertype)
        self._dropdowns.append(widget)
        return widget
 
    def createsearchfilter(self, filtertype="MFG", columnname="SUPP_MFG"):
        sf_font = max(7, self._sz(9))
        btn_size = max(16, self._sz(22))
        widget = QWidget()
        widget.setMinimumHeight(max(40, self._sz(70)))
        widget.setMinimumWidth(max(100, self._sz(150)))

        layout = QVBoxLayout()
        layout.setContentsMargins(3, 2, 3, 2)

        label = QLabel(f"{filtertype} Search")
        label.setFont(QFont("Arial", sf_font, QFont.Bold))
        layout.addWidget(label)

        inputrow = QHBoxLayout()
        inputrow.setSpacing(2)

        searchinput = QLineEdit()
        searchinput.setPlaceholderText(f"Enter {filtertype}...")
        searchinput.setStyleSheet(f"QLineEdit {{border: 2px solid #ccc; border-radius: 3px; padding: 3px; font-size: {sf_font}px;}} QLineEdit:focus {{border-color: #4CAF50;}}")
        inputrow.addWidget(searchinput, 1)

        searchbtn = QPushButton("✓")
        searchbtn.setFixedSize(btn_size, btn_size)
        searchbtn.clicked.connect(self.applyfilters)
        searchbtn.setStyleSheet("QPushButton {background-color: #156082; color: white; border: none; padding: 1px; border-radius: 3px;} QPushButton:hover {background-color: #45a049;}")
        inputrow.addWidget(searchbtn)

        clearsearchbtn = QPushButton("✖")
        clearsearchbtn.setFixedSize(btn_size, btn_size)
        clearsearchbtn.clicked.connect(lambda: self.clearsearchfilter(widget))
        clearsearchbtn.setStyleSheet("QPushButton {background-color: #E97132; color: white; border: none; padding: 1px; border-radius: 3px;} QPushButton:hover {background-color: #da190b;}")
        inputrow.addWidget(clearsearchbtn)

        layout.addLayout(inputrow)

        statuslabel = QLabel("")
        statuslabel.setStyleSheet(f"color: #666; font-size: {sf_font}px;")
        statuslabel.setWordWrap(True)
        layout.addWidget(statuslabel)

        layout.addStretch()

        searchinput.returnPressed.connect(self.applyfilters)

        widget.searchinput = searchinput
        widget.statuslabel = statuslabel
        widget.filtertype = filtertype
        widget.columnname = columnname

        widget.setLayout(layout)
        return widget
 
    def clearsearchfilter(self, searchwidget):
        searchwidget.searchinput.clear()
        searchwidget.statuslabel.setText("")
        self.applyfilters()
 
    def populatefilters(self, coveragedf):
        if 'SCC Name' in coveragedf.columns:
            uniquescc = coveragedf['SCC Name'].dropna().unique()
            self.sccfilter.additems(uniquescc)
 
        if 'Region' in coveragedf.columns:
            uniqueregions = coveragedf['Region'].dropna().unique()
            self.regionfilter.additems(uniqueregions)
 
        if 'Program Supported' in coveragedf.columns:
            uniqueprograms = coveragedf['Program Supported'].dropna().unique()
            self.programsupportedfilter.additems(uniqueprograms)

        if 'Day Alert' in coveragedf.columns:
            uniqueday = coveragedf['Day Alert'].dropna().unique()
            validdays = []
            for d in uniqueday:
                try:
                    validdays.append(int(d))
                except (ValueError, TypeError):
                    continue
 
            dayoptions = []
            for dayint in sorted(validdays):
                if dayint == 0:
                    dayoptions.append("Day 0")
                elif dayint >= 999:
                    dayoptions.append("Covered")
                else:
                    dayoptions.append(f"Day {dayint}")
 
            self.dayalertfilter.additems(dayoptions, presorted=True)
 
    def applyfilters(self):
        if not hasattr(self, 'originalcoveragedf'):
            return
 
        self.coveragetable.setEnabled(False)
 
        try:
            filtereddf = self.originalcoveragedf.copy()
 
            selectedscc = self.sccfilter.getselecteditems()
            if selectedscc and 'SCC Name' in filtereddf.columns:
                filtereddf = filtereddf[filtereddf['SCC Name'].isin(selectedscc)]
 
            selectedregion = self.regionfilter.getselecteditems()
            if selectedregion and 'Region' in filtereddf.columns:
                filtereddf = filtereddf[filtereddf['Region'].isin(selectedregion)]
 
            selectedprograms = self.programsupportedfilter.getselecteditems()
            if selectedprograms and 'Program Supported' in filtereddf.columns:
                filtereddf = filtereddf[filtereddf['Program Supported'].isin(selectedprograms)]

            selecteddays = self.dayalertfilter.getselecteditems()
            if selecteddays and 'Day Alert' in filtereddf.columns:
                dayvalues = []
                for daystr in selecteddays:
                    if "Day 0" in daystr:
                        dayvalues.append(0)
                    elif "Covered" in daystr:
                        dayvalues.append(999)
                    elif "Day " in daystr:
                        try:
                            dayvalues.append(int(daystr.split()[1]))
                        except (ValueError, IndexError):
                            continue
                if dayvalues:
                    filtereddf = filtereddf[filtereddf['Day Alert'].isin(dayvalues)]

            if hasattr(self, '_partlist_textedit'):
                raw = self._partlist_textedit.toPlainText().strip()
                if raw and 'Part Number' in filtereddf.columns:
                    normalized = raw.replace(',', '\n').replace(';', '\n').replace('\t', '\n')
                    partnums = {p.strip().upper() for p in normalized.splitlines() if p.strip()}
                    if partnums:
                        filtereddf = filtereddf[
                            filtereddf['Part Number'].astype(str).str.upper().isin(partnums)
                        ]
                        self._partlist_statuslabel.setText(
                            f"{len(filtereddf)} of {len(partnums)} found"
                        )
                    else:
                        self._partlist_statuslabel.setText("")
                else:
                    self._partlist_statuslabel.setText("")

            searchcolumnmapping = {
                'MFG': 'MFG Code',
                'Part': 'Part Number',
                'SHP': 'SHP Code',
                'Country': 'SHP Country',
            }
 
            for searchfilter in self.searchfilters:
                searchtext = searchfilter.searchinput.text().strip().upper()
                filtertype = searchfilter.filtertype
                columnname = searchcolumnmapping.get(filtertype, searchfilter.columnname)
 
                if searchtext and columnname in filtereddf.columns:
                    if ',' in searchtext:
                        codes = [code.strip() for code in searchtext.split(',') if code.strip()]
                        filtereddf = filtereddf[
                            filtereddf[columnname].astype(str).str.upper().isin(codes)
                        ]
                        searchfilter.statuslabel.setText(f"Searching: {', '.join(codes)}")
                    else:
                        filtereddf = filtereddf[
                            filtereddf[columnname].astype(str).str.upper().str.contains(
                                searchtext, na=False
                            )
                        ]
                        searchfilter.statuslabel.setText(f"Searching: {searchtext}")
                else:
                    searchfilter.statuslabel.setText("")
 
            max_display_rows = 2000
            if len(filtereddf) > max_display_rows:
                filtereddf = filtereddf.head(max_display_rows)

            self.currentcoveragedf = filtereddf.copy()
 
            self.displaycoveragetable(filtereddf)
 
        finally:
            self.coveragetable.setEnabled(True)
 
    def clearfilters(self):
        if hasattr(self, 'sccfilter'):
            self.sccfilter.selectallitems()
        if hasattr(self, 'regionfilter'):
            self.regionfilter.selectallitems()
        if hasattr(self, 'dayalertfilter'):
            self.dayalertfilter.selectallitems()
        if hasattr(self, 'programsupportedfilter'):
            self.programsupportedfilter.selectallitems()
        if hasattr(self, 'searchfilters'):
            for searchfilter in self.searchfilters:
                searchfilter.searchinput.clear()
                searchfilter.statuslabel.setText("")
        self._clearpartlistfilter(refilter=False)

        if hasattr(self, 'originalcoveragedf'):
            self.currentcoveragedf = self.originalcoveragedf.copy()
            self.displaycoveragetable(self.currentcoveragedf)

    def _clearpartlistfilter(self, refilter=True):
        if hasattr(self, '_partlist_timer'):
            self._partlist_timer.stop()
        if hasattr(self, '_partlist_textedit'):
            self._partlist_textedit.blockSignals(True)
            self._partlist_textedit.clear()
            self._partlist_textedit.blockSignals(False)
        if hasattr(self, '_partlist_statuslabel'):
            self._partlist_statuslabel.setText("")
        if refilter:
            self.applyfilters()
 
    def createcoveragedashboard(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("Coverage Dashboard")
        title.setFont(QFont("Arial", max(10, self._sz(16)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        buttonlayout = QHBoxLayout()

        _bpv = max(4, self._sz(10))
        _bph = max(8, self._sz(20))
        loadbtn = QPushButton("Generate Coverage Analysis")
        loadbtn.clicked.connect(self.generatecoverageanalysis)
        loadbtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #45a049;}}")
        buttonlayout.addWidget(loadbtn)

        refreshbtn = QPushButton("Refresh Data")
        refreshbtn.clicked.connect(self.refreshcoveragedata)
        buttonlayout.addWidget(refreshbtn)

        exportbtn = QPushButton("Export to XLSX")
        exportbtn.clicked.connect(self.exportcoveragetable)
        buttonlayout.addWidget(exportbtn)

        columnsbtn = QPushButton("Hide Columns")
        columnsbtn.clicked.connect(self.showcoveragecolumnmenu)
        buttonlayout.addWidget(columnsbtn)
        self._coverage_column_menu = columnsbtn

        resetsortbtn = QPushButton("Reset Sort")
        resetsortbtn.clicked.connect(self.resetcoveragetablesort)
        buttonlayout.addWidget(resetsortbtn)

        freezebtn = QPushButton("Freeze Columns")
        freezebtn.clicked.connect(self.showfreezecolumnmenu)
        buttonlayout.addWidget(freezebtn)
        self._freeze_cols_btn = freezebtn

        uploadcommentsbtn = QPushButton("Upload Comments")
        uploadcommentsbtn.clicked.connect(self._uploadcoveragecomments)
        uploadcommentsbtn.setStyleSheet(f"QPushButton {{background-color: #D17000; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #E8860A;}}")
        buttonlayout.addWidget(uploadcommentsbtn)

        buttonlayout.addStretch()
        layout.addLayout(buttonlayout)

        self.filtersection = self.createfiltersection()
        layout.addWidget(self.filtersection)

        self._frozen_cols = set()
        self.coveragetable = QTableWidget()
        self.coveragetable.setWordWrap(True)
        self.coveragetable.setSortingEnabled(True)
        self.coveragetable.horizontalHeader().sectionResized.connect(
            lambda _col, _old, _new: self._update_frozen_geometry())
        self.coveragetable.horizontalScrollBar().valueChanged.connect(
            lambda _: (
                self.coveragetable.viewport().update(),
                self.coveragetable.horizontalHeader().viewport().update()
            ) if self._frozen_cols else None)
        layout.addWidget(self.coveragetable)

        self._frozen_view = QTableView(self.coveragetable)
        self._frozen_view.setFocusPolicy(Qt.NoFocus)
        self._frozen_view.verticalHeader().hide()
        self._frozen_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._frozen_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._frozen_view.setStyleSheet(
            "QTableView { border: none; }"
            "QHeaderView::section { background-color: #e8e8e8;"
            " border: 1px solid #d0d0d0; font-weight: bold; }"
        )
        self._frozen_view.hide()
        self._frozen_view.viewport().installEventFilter(self)
        self.coveragetable.installEventFilter(self)

        widget.setLayout(layout)
        return widget

    def showcoveragecolumnmenu(self):
        if self.coveragetable.columnCount() == 0:
            return
        
        menu = PersistentMenu(self)

        showallaction = QAction("Show All Columns", self)
        showallaction.triggered.connect(self.showallcoveragecolumns)
        menu.addAction(showallaction)

        menu.addSeparator()

        for col in range(self.coveragetable.columnCount()):
            headeritem = self.coveragetable.horizontalHeaderItem(col)
            if not headeritem:
                continue
            
            colname = headeritem.text()
            action = QAction(colname, self)
            action.setCheckable(True)
            action.setChecked(colname not in self._hidden_coverage_columns)
            action.toggled.connect(lambda checked, name=colname: self.setcoveragecolumnvisible(name, checked))
            menu.addAction(action)
        
        menu.exec_(self._coverage_column_menu.mapToGlobal(self._coverage_column_menu.rect().bottomLeft()))

    def showfreezecolumnmenu(self):
        if self.coveragetable.columnCount() == 0:
            return
        menu = PersistentMenu(self)

        clearaction = QAction("Unfreeze All", self)
        clearaction.triggered.connect(self._clearfrozencolumns)
        menu.addAction(clearaction)
        menu.addSeparator()

        for col in range(self.coveragetable.columnCount()):
            header = self.coveragetable.horizontalHeaderItem(col)
            if not header or self.coveragetable.isColumnHidden(col):
                continue
            name = header.text()
            action = QAction(name, self)
            action.setCheckable(True)
            action.setChecked(name in self._frozen_cols)
            action.toggled.connect(lambda checked, n=name: self._togglefreezecolumn(n, checked))
            menu.addAction(action)

        menu.exec_(self._freeze_cols_btn.mapToGlobal(self._freeze_cols_btn.rect().bottomLeft()))

    def _togglefreezecolumn(self, colname, freeze):
        if freeze:
            self._frozen_cols.add(colname)
        else:
            self._frozen_cols.discard(colname)
        self._applyfrozencolumns()

    def _clearfrozencolumns(self):
        self._frozen_cols.clear()
        self._frozen_view.hide()

    def setcoveragecolumnvisible(self, columnname, visible):
        colindex = self._getcoveragecolumnindex(columnname)
        if colindex is None:
            return
        self.coveragetable.setColumnHidden(colindex, not visible)
        if visible:
            self._hidden_coverage_columns.discard(columnname)
        else:
            self._hidden_coverage_columns.add(columnname)
        self._applyfrozencolumns()

    def showallcoveragecolumns(self):
        if self.coveragetable.columnCount() == 0:
            return
        self._hidden_coverage_columns.clear()
        for col in range(self.coveragetable.columnCount()):
            self.coveragetable.setColumnHidden(col, False)
        self._applyfrozencolumns()

    def _getcoveragecolumnindex(self, columnname):
        for col in range(self.coveragetable.columnCount()):
            item = self.coveragetable.horizontalHeaderItem(col)
            if item and item.text() == columnname:
                return col
        return None

    def _reapplycoveragehiddencolumns(self):
        if self.coveragetable.columnCount() == 0:
            return
        for col in range(self.coveragetable.columnCount()):
            item = self.coveragetable.horizontalHeaderItem(col)
            if not item:
                continue
            self.coveragetable.setColumnHidden(col, item.text() in self._hidden_coverage_columns)

    def _applyfrozencolumns(self):
        ct = self.coveragetable
        fv = self._frozen_view
        if fv.model() is not ct.model():
            fv.setModel(ct.model())
            # QAbstractItemDelegate.commitData is the signal (not QAbstractItemView.commitData
            # which is a slot). QueuedConnection ensures our handler runs after setModelData
            # has written the new value to the model — direct connection would fire first,
            # before the item text is updated.
            fv.itemDelegate().commitData.connect(
                self._on_frozen_view_commit, Qt.QueuedConnection
            )
            # Bidirectional column-width sync: main table → frozen view is handled
            # inside _update_frozen_geometry; this reverse leg syncs a drag in the
            # frozen overlay back to the main table, which then re-fires sectionResized
            # and triggers _update_frozen_geometry so the overlay width grows/shrinks.
            fv.horizontalHeader().sectionResized.connect(self._on_frozen_column_resized)
            ct.verticalScrollBar().valueChanged.connect(fv.verticalScrollBar().setValue)
            fv.verticalScrollBar().valueChanged.connect(ct.verticalScrollBar().setValue)
            ct.verticalHeader().sectionResized.connect(
                lambda idx, _old, new: fv.verticalHeader().resizeSection(idx, new)
            )
        ncols = ct.columnCount()
        if not self._frozen_cols or ncols == 0:
            fv.hide()
            return
        # Block fv's header signals while configuring column visibility so that
        # setColumnHidden(c, True) doesn't emit sectionResized(c, old, 0) and
        # trigger _on_frozen_column_resized, which would set ct.columnWidth(c)=0
        # and make every unfrozen column invisible in the main table.
        fv.horizontalHeader().blockSignals(True)
        for c in range(ncols):
            header = ct.horizontalHeaderItem(c)
            name = header.text() if header else ''
            frozen = name in self._frozen_cols and not ct.isColumnHidden(c)
            fv.setColumnHidden(c, not frozen)
            if frozen:
                fv.setColumnWidth(c, ct.columnWidth(c))
        fv.horizontalHeader().blockSignals(False)
        for r in range(ct.rowCount()):
            fv.verticalHeader().resizeSection(r, ct.rowHeight(r))
        self._update_frozen_geometry()
        # Reset scroll first so frozen columns sit at the left edge of the viewport.
        ct.horizontalScrollBar().setValue(0)
        # Repaint the main table synchronously BEFORE showing the overlay so
        # the unfrozen column area is fully rendered at position 0 before the
        # frozen view appears on top.  Showing the overlay first and repainting
        # afterward leaves a window where Qt can clear the unfrozen area.
        ct.viewport().repaint()
        ct.horizontalHeader().viewport().repaint()
        fv.show()
        fv.raise_()
        # Deferred second pass to catch any paint artefacts introduced by the
        # overlay widget becoming visible in the hierarchy.
        QTimer.singleShot(0, ct.viewport().update)
        QTimer.singleShot(0, ct.horizontalHeader().viewport().update)

    def _on_frozen_view_commit(self, _editor):
        idx = self._frozen_view.currentIndex()
        if not idx.isValid():
            return
        item = self.coveragetable.item(idx.row(), idx.column())
        if item:
            self.oncommentchanged(item)

    def _on_frozen_column_resized(self, col, _old_size, new_size):
        ct = self.coveragetable
        if ct.columnWidth(col) == new_size:
            return
        # Block ct's header signals so setColumnWidth doesn't synchronously fire
        # sectionResized → _update_frozen_geometry → fv.setGeometry() mid-drag.
        # A mid-drag setGeometry on the frozen view cancels the drag operation.
        ct.horizontalHeader().blockSignals(True)
        ct.setColumnWidth(col, new_size)
        ct.horizontalHeader().blockSignals(False)
        # Defer geometry update to the next event loop tick — by then the drag
        # step is complete and resizing the overlay won't interrupt anything.
        QTimer.singleShot(0, self._update_frozen_geometry)

    def _update_frozen_geometry(self):
        ct = self.coveragetable
        fv = self._frozen_view
        if not self._frozen_cols or ct.columnCount() == 0:
            return
        frozen_width = 0
        for c in range(ct.columnCount()):
            header = ct.horizontalHeaderItem(c)
            name = header.text() if header else ''
            if name in self._frozen_cols and not ct.isColumnHidden(c):
                fv.setColumnWidth(c, ct.columnWidth(c))
                frozen_width += ct.columnWidth(c)
        if frozen_width == 0:
            fv.hide()
            return
        vhw = ct.verticalHeader().width()
        fw = ct.frameWidth()
        hh = ct.horizontalHeader().height()
        fv.horizontalHeader().setFixedHeight(hh)
        fv.setGeometry(vhw + fw, fw, frozen_width, ct.viewport().height() + hh)
        # After resizing the overlay (especially when contracting), Qt does not
        # automatically repaint the now-exposed area of the main viewport.
        # Schedule a deferred update so revealed columns render correctly.
        ct.viewport().update()
        ct.horizontalHeader().viewport().update()

    def eventFilter(self, obj, event):
        if obj is self._frozen_view and event.type() == QEvent.Wheel:
            QApplication.sendEvent(self.coveragetable.viewport(), event)
            return True
        if obj is self.coveragetable and event.type() == QEvent.Resize:
            self._update_frozen_geometry()
        return super().eventFilter(obj, event)

    def createcalloffforecasttab(self):
        widget = QWidget()
        layout = QVBoxLayout()
 
        title = QLabel("Call-off Forecast and Waterfall")
        title.setFont(QFont("Arial", max(10, self._sz(16)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
 
        searchsection = self.createcalloffsearchsection()
        layout.addWidget(searchsection)
 
        self.dailyviewsection = self.createdailyviewsection()
        layout.addWidget(self.dailyviewsection)
 
        widget.setLayout(layout)
        return widget
 
    def createindividualpartcoverage(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("Individual Part Coverage")
        title.setFont(QFont("Arial", max(10, self._sz(16)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        searchsection = self.createpartsearchsection()
        layout.addWidget(searchsection)

        inforow = QHBoxLayout()
        self.partinfosection = self.createpartinfosection()
        inforow.addWidget(self.partinfosection, stretch=2)

        self.partnotessection = self._createpartnotessection()
        inforow.addWidget(self.partnotessection, stretch=1)
        layout.addLayout(inforow)

        self.transactionsection = self.createtransactionsection()
        layout.addWidget(self.transactionsection)

        widget.setLayout(layout)
        return widget
 
    def createldjiscoveragetab(self):
        widget = QWidget()
        layout = QVBoxLayout()
 
        title = QLabel("LDJIS Coverage")
        title.setFont(QFont("Arial", max(10, self._sz(16)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        _bpv = max(4, self._sz(10))
        _bph = max(8, self._sz(20))
        btnlayout = QHBoxLayout()
        generatebtn = QPushButton("Generate LDJIS Coverage")
        generatebtn.clicked.connect(self.generateldjiscoverage)
        generatebtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #45a049;}}")
        btnlayout.addWidget(generatebtn)
        btnlayout.addStretch()
        layout.addLayout(btnlayout)
 
        tablestyle = """QTableWidget {gridline-color: #d0d0d0; background-color: white;} QTableWidget::item {padding: 6px; border: 1px solid #d0d0d0;} QHeaderView::section {background-color: #f0f0f0; padding: 8px; border: 1px solid #d0d0d0; font-weight: bold;}"""
 
        scrollarea = QScrollArea()
        scrollarea.setWidgetResizable(True)
        self.ldjistable = QTableWidget()
        self.ldjistable.setWordWrap(True)
        self.ldjistable.setSortingEnabled(False)
        self.ldjistable.setStyleSheet(tablestyle)
        scrollarea.setWidget(self.ldjistable)
        layout.addWidget(scrollarea)
 
        widget.setLayout(layout)
        return widget
 
    def createalertstab(self):
        widget = QWidget()
        layout = QVBoxLayout()
 
        title = QLabel("Alerts Breakdown")
        title.setFont(QFont("Arial", max(10, self._sz(16)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        _bpv = max(4, self._sz(10))
        _bph = max(8, self._sz(20))
        btnlayout = QHBoxLayout()
        generatebtn = QPushButton("Generate Alerts Breakdown")
        generatebtn.clicked.connect(self.generatealertsbreakdown)
        generatebtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #45a049;}}")
        btnlayout.addWidget(generatebtn)
 
        exportbtn = QPushButton("Export to XLSX")
        exportbtn.clicked.connect(self.exportalertstable)
        btnlayout.addWidget(exportbtn)

        uploadfieldsbtn = QPushButton("Upload Field Data")
        uploadfieldsbtn.clicked.connect(self._uploadalertsdata)
        uploadfieldsbtn.setStyleSheet("""QPushButton {background-color: #D17000; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;} QPushButton:hover {background-color: #E8860A;}""")
        btnlayout.addWidget(uploadfieldsbtn)

        btnlayout.addStretch()
        layout.addLayout(btnlayout)
 
        self.alertfiltersection = self.createalertfiltersection()
        layout.addWidget(self.alertfiltersection)
 
        _tpad = max(2, self._sz(4))
        _hpad = max(4, self._sz(6))
        _tfont = max(9, self._sz(11))
        tablestyle = f"QTableWidget {{gridline-color: #d0d0d0; background-color: white; font-size: {_tfont}px;}} QTableWidget::item {{padding: {_tpad}px; border: 1px solid #d0d0d0;}} QHeaderView::section {{background-color: #f0f0f0; padding: {_hpad}px; border: 1px solid #d0d0d0; font-weight: bold; font-size: {_tfont}px;}}"

        scrollarea = QScrollArea()
        scrollarea.setWidgetResizable(True)
        self.alertstable = QTableWidget()
        self.alertstable.setWordWrap(True)
        self.alertstable.setSortingEnabled(True)
        self.alertstable.setStyleSheet(tablestyle)
        self.alertstable.setItemDelegate(_BackgroundDelegate(self.alertstable))
        if self.userdata.get('username', '') in ADMINUSERS:
            self.alertstable.setContextMenuPolicy(Qt.CustomContextMenu)
            self.alertstable.customContextMenuRequested.connect(self._onalerts_contextmenu)
        scrollarea.setWidget(self.alertstable)
        layout.addWidget(scrollarea)
 
        widget.setLayout(layout)
        return widget
    
    def createpiwdtab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("PIWD Report")
        title.setFont(QFont("Arial", max(10, self._sz(15)), QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        _bpv = max(4, self._sz(10))
        _bph = max(8, self._sz(20))
        btnlayout = QHBoxLayout()
        generatebtn = QPushButton("Generate PIWD Report")
        generatebtn.clicked.connect(self.generatepiwdreport)
        generatebtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #45a049;}}")
        btnlayout.addWidget(generatebtn)
        
        exportbtn = QPushButton("Export to XLSX")
        exportbtn.clicked.connect(self.exportpiwdreport)
        btnlayout.addWidget(exportbtn)
        
        btnlayout.addStretch()
        layout.addLayout(btnlayout)

        self.piwdfiltersection = self.createpiwdfiltersection()
        layout.addWidget(self.piwdfiltersection)
        
        _tpad = max(2, self._sz(4))
        _hpad = max(4, self._sz(6))
        _tfont = max(9, self._sz(11))
        tablestyle = f"QTableWidget {{gridline-color: #d0d0d0; background-color: white; font-size: {_tfont}px;}} QTableWidget::item {{padding: {_tpad}px; border: 1px solid #d0d0d0;}} QHeaderView::section {{background-color: #f0f0f0; padding: {_hpad}px; border: 1px solid #d0d0d0; font-weight: bold; font-size: {_tfont}px;}}"

        scrollarea = QScrollArea()
        scrollarea.setWidgetResizable(True)
        self.piwdtable = QTableWidget()
        self.piwdtable.setWordWrap(True)
        self.piwdtable.setSortingEnabled(True)
        self.piwdtable.setStyleSheet(tablestyle)
        scrollarea.setWidget(self.piwdtable)
        layout.addWidget(scrollarea)
        
        widget.setLayout(layout)
        return widget
 
    def createpiwdfiltersection(self):
        widget = QWidget()
        widget.setMaximumHeight(self._alert_filter_h)

        layout = QHBoxLayout()

        filterlabel = QLabel("Filters:")
        filterlabel.setFont(QFont("Arial", max(8, self._sz(12)), QFont.Bold))
        filterlabel.setAlignment(Qt.AlignTop)
        layout.addWidget(filterlabel)

        self.piwd_scc_filter = self.createmultiselectdropdown("Select SCC Names...", "SCC")
        self.piwd_scc_filter.selectionChanged.connect(self.applypiwdfilters)
        layout.addWidget(self.piwd_scc_filter)

        self.piwd_part_filter = self.createmultiselectdropdown("Select Parts...", "Part")
        self.piwd_part_filter.selectionChanged.connect(self.applypiwdfilters)
        layout.addWidget(self.piwd_part_filter)

        clearfiltersbtn = QPushButton("Clear All Filters")
        clearfiltersbtn.clicked.connect(self.clearpiwdfilters)
        clearfiltersbtn.setMaximumWidth(120)
        clearfiltersbtn.setMaximumHeight(30)
        layout.addWidget(clearfiltersbtn)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def createcalloffsearchsection(self):
        widget = QWidget()
        widget.setMaximumHeight(max(40, self._sz(65)))
        _sf = max(8, self._sz(12))
        _bpv = max(4, self._sz(8))
        _bph = max(8, self._sz(16))

        layout = QHBoxLayout()

        searchlabel = QLabel("Part Number:")
        searchlabel.setFont(QFont("Arial", _sf, QFont.Bold))
        layout.addWidget(searchlabel)

        self.calloffpartsearch = QLineEdit()
        self.calloffpartsearch.setPlaceholderText("Enter part number for call-off analysis...")
        self.calloffpartsearch.setMaximumWidth(self._sz(250))
        self.calloffpartsearch.setStyleSheet(f"QLineEdit {{border: 2px solid #ccc; border-radius: 5px; padding: {_bpv}px; font-size: {_sf}px;}} QLineEdit:focus {{border-color: #4CAF50;}}")
        self.calloffpartsearch.returnPressed.connect(self.searchcalloffdata)
        layout.addWidget(self.calloffpartsearch)

        searchbtn = QPushButton("Generate Waterfall")
        searchbtn.clicked.connect(self.searchcalloffdata)
        searchbtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #1976D2;}}")
        layout.addWidget(searchbtn)

        clearbtn = QPushButton("Clear")
        clearbtn.clicked.connect(self.clearcalloffanalysis)
        clearbtn.setStyleSheet(f"QPushButton {{background-color: #E97132; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #da190b;}}")
        layout.addWidget(clearbtn)
 
        layout.addStretch()
        widget.setLayout(layout)
        return widget
 
    def createdailyviewsection(self):
        widget = QWidget()
        layout = QVBoxLayout()
        tablestyle = """QTableWidget {gridline-color: #d0d0d0; background-color: white;} QTableWidget::item {padding: 8px; border: 1px solid #d0d0d0;} QHeaderView::section {background-color: #f0f0f0; padding: 10px; border: 1px solid #d0d0d0; font-weight: bold;}"""
 
        dailytitlerow = QHBoxLayout()
        dailytitle = QLabel("Daily View - Call-off Forecast Waterfall")
        dailytitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        dailytitlerow.addWidget(dailytitle)
        dailydeltatitle = QLabel("Daily Delta")
        dailydeltatitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        dailytitlerow.addWidget(dailydeltatitle)
        layout.addLayout(dailytitlerow)
 
        dailytablesrow = QHBoxLayout()
 
        dailyscroll = QScrollArea()
        dailyscroll.setWidgetResizable(True)
        self.dailyviewtable = QTableWidget()
        self.dailyviewtable.setSortingEnabled(False)
        self.dailyviewtable.setStyleSheet(tablestyle)
        dailyscroll.setWidget(self.dailyviewtable)
        dailytablesrow.addWidget(dailyscroll)
 
        dailydeltascroll = QScrollArea()
        dailydeltascroll.setWidgetResizable(True)
        self.dailydeltatable = QTableWidget()
        self.dailydeltatable.setSortingEnabled(False)
        self.dailydeltatable.setStyleSheet(tablestyle)
        dailydeltascroll.setWidget(self.dailydeltatable)
        dailytablesrow.addWidget(dailydeltascroll)
 
        layout.addLayout(dailytablesrow)
 
        weeklytitlerow = QHBoxLayout()
        weeklytitle = QLabel("Weekly Summary - Call-off Forecast Waterfall")
        weeklytitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        weeklytitlerow.addWidget(weeklytitle)
        weeklydeltatitle = QLabel("Weekly Delta")
        weeklydeltatitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        weeklytitlerow.addWidget(weeklydeltatitle)
        layout.addLayout(weeklytitlerow)
 
        weeklytablesrow = QHBoxLayout()
 
        weeklyscroll = QScrollArea()
        weeklyscroll.setWidgetResizable(True)
        self.weeklyviewtable = QTableWidget()
        self.weeklyviewtable.setSortingEnabled(False)
        self.weeklyviewtable.setStyleSheet(tablestyle)
        weeklyscroll.setWidget(self.weeklyviewtable)
        weeklytablesrow.addWidget(weeklyscroll)
 
        weeklydeltascroll = QScrollArea()
        weeklydeltascroll.setWidgetResizable(True)
        self.weeklydeltatable = QTableWidget()
        self.weeklydeltatable.setSortingEnabled(False)
        self.weeklydeltatable.setStyleSheet(tablestyle)
        weeklydeltascroll.setWidget(self.weeklydeltatable)
        weeklytablesrow.addWidget(weeklydeltascroll)
 
        layout.addLayout(weeklytablesrow)
 
        widget.setLayout(layout)
        return widget
 
    def createpartsearchsection(self):
        widget = QWidget()
        widget.setMaximumHeight(max(40, self._sz(65)))
        _sf = max(8, self._sz(12))
        _bpv = max(4, self._sz(8))
        _bph = max(8, self._sz(16))
        layout = QHBoxLayout()

        searchlabel = QLabel("Part Number:")
        searchlabel.setFont(QFont("Arial", _sf, QFont.Bold))
        layout.addWidget(searchlabel)

        self.partnumbersearch = QLineEdit()
        self.partnumbersearch.setPlaceholderText("Enter part number...")
        self.partnumbersearch.setMaximumWidth(self._sz(200))
        self.partnumbersearch.setStyleSheet(f"QLineEdit {{border: 2px solid #ccc; border-radius: 5px; padding: {_bpv}px; font-size: {_sf}px;}} QLineEdit:focus {{border-color: #4CAF50;}}")
        self.partnumbersearch.returnPressed.connect(self.searchpartcoverage)
        layout.addWidget(self.partnumbersearch)

        searchbtn = QPushButton("Search")
        searchbtn.clicked.connect(self.searchpartcoverage)
        searchbtn.setStyleSheet(f"QPushButton {{background-color: #156082; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #45a049;}}")
        layout.addWidget(searchbtn)

        clearbtn = QPushButton("Clear")
        clearbtn.clicked.connect(self.clearpartsearch)
        clearbtn.setStyleSheet(f"QPushButton {{background-color: #E97132; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #da190b;}}")
        layout.addWidget(clearbtn)

        exportbtn = QPushButton("Export Transactions")
        exportbtn.clicked.connect(self.exporttransactiontable)
        exportbtn.setStyleSheet(f"QPushButton {{background-color: #1976D2; color: white; padding: {_bpv}px {_bph}px; border: none; border-radius: 5px; font-weight: bold;}} QPushButton:hover {{background-color: #1565C0;}}")
        layout.addWidget(exportbtn)
 
        layout.addStretch()
        widget.setLayout(layout)
        return widget
 
    def createpartinfosection(self):
        widget = QWidget()
 
        layout = QVBoxLayout()
        infotitle = QLabel("Part Information")
        infotitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        layout.addWidget(infotitle)

        gridwidget = QWidget()
        gridlayout = QGridLayout(gridwidget)
        gridlayout.setSpacing(self._sz(10))
 
        self.partinfolabels = {}
        infofields = [
            ("Part Description", 0, 0), ("Supplier Name", 0, 1), ("MFG Code", 0, 2), ("SHP Code", 0, 3),
            ("Unit Load Qty", 1, 0), ("Multi Unit Load", 1, 1), ("Piece Price", 1, 2),
            ("Safety Stock", 2, 0), ("Safety Days", 2, 1), ("Initial Stock", 2, 2),
        ]
 
        for fieldname, row, col in infofields:
            label = QLabel(f"{fieldname}:")
            label.setFont(QFont("Arial", max(7, self._sz(10)), QFont.Bold))
            gridlayout.addWidget(label, row * 2, col)
 
            valuelabel = QLabel("--")
            valuelabel.setStyleSheet(f"color: #333; font-size: {max(7, self._sz(11))}px; padding: {self._sz(2)}px;")
            gridlayout.addWidget(valuelabel, row * 2 + 1, col)
 
            self.partinfolabels[fieldname] = valuelabel
 
        layout.addWidget(gridwidget)
        widget.setLayout(layout)
        return widget
 
    def createtransactionsection(self):
        widget = QWidget()
        layout = QVBoxLayout()
 
        transtitle = QLabel("Transaction Projections")
        transtitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        layout.addWidget(transtitle)
 
        scrollarea = QScrollArea()
        scrollarea.setWidgetResizable(True)
 
        self.transactiontable = QTableWidget()
        self.transactiontable.setSortingEnabled(False)
 
        columnheaders = ["Date", "Transaction Type", "Receipt/Reqmt", "Available QTY", "ASN"]
        self.transactiontable.setColumnCount(len(columnheaders))
        self.transactiontable.setHorizontalHeaderLabels(columnheaders)
 
        self.transactiontable.setColumnWidth(0, self._sz(100))
        self.transactiontable.setColumnWidth(1, self._sz(120))
        self.transactiontable.setColumnWidth(2, self._sz(110))
        self.transactiontable.setColumnWidth(3, self._sz(110))
        self.transactiontable.setColumnWidth(4, self._sz(100))
 
        scrollarea.setWidget(self.transactiontable)
        layout.addWidget(scrollarea)
 
        widget.setLayout(layout)
        return widget
 
    def searchpartcoverage(self):
        partnumber = self.partnumbersearch.text().strip().upper()
 
        if not partnumber:
            QMessageBox.warning(self, "Missing Input", "Please enter a part number.")
            return
 
        try:
            success, message, datadict = self.coverageengine.loadrequireddata()
            if not success:
                QMessageBox.warning(self, "Missing Data", message)
                return
 
            partinfo, transactions = self.coverageengine.analyzeindivpart(partnumber, datadict)
            if not partinfo:
                QMessageBox.information(self, "Part Not Found", f"Part number {partnumber} was not found in the system.")
                return
 
            self._currentpartnumber = partnumber
            self.displaypartinfo(partinfo)
            self.displaytransactiontable(transactions)
            self._loadnoteforpart(partnumber)
 
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Search failed: {str(e)}")
 
    def searchcalloffdata(self):
        partnumber = self.calloffpartsearch.text().strip().upper()
 
        if not partnumber:
            QMessageBox.warning(self, "Missing Input", "Please enter a part number.")
            return
 
        try:
            valid, message, dates = self.waterfallengine.validatearchiveavailability()
            if not valid:
                QMessageBox.warning(self, "Archive Not Found", f"{message}\n\nExpected directory:\n{self.waterfallengine.archive_dir}")
                return
 
            waterfalldata = self.waterfallengine.generatecalloffwaterfall(partnumber)
 
            if not waterfalldata:
                QMessageBox.information(self, "No Data", f"No call-off data found for part {partnumber}.")
                return
 
            shippingdates = self.waterfallengine.generateshippingdaterange(datetime.now().date(), 90)
            self.displaydailyviewtable(waterfalldata)
            self.displaydailydeltatable(waterfalldata, shippingdates)
            self.displayweeklyviewtable(waterfalldata, shippingdates)
            self.displayweeklydeltatable(waterfalldata, shippingdates)
 
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Call-off analysis failed: {str(e)}")
 
    def displaydailyviewtable(self, waterfalldata):
        if not waterfalldata:
            return
 
        today = datetime.now().date()
        shippingdates = self.waterfallengine.generateshippingdaterange(today, 90)
        headers = ['Call off Date'] + [date.strftime('%m/%d') for date in shippingdates]
 
        self.dailyviewtable.setRowCount(len(waterfalldata))
        self.dailyviewtable.setColumnCount(len(headers))
        self.dailyviewtable.setHorizontalHeaderLabels(headers)
        self.dailyviewtable.setColumnWidth(0, 120)
 
        for col in range(1, len(headers)):
            self.dailyviewtable.setColumnWidth(col, 80)
 
        for rowindex, rowdata in enumerate(waterfalldata):
            archivedate = rowdata['archive_date']
            calloffs = rowdata['call_offs']
            filefound = rowdata.get('file_found', True)
 
            dateitem = QTableWidgetItem(archivedate.strftime('%m/%d'))
            dateitem.setFlags(dateitem.flags() & ~Qt.ItemIsEditable)
 
            if filefound:
                dateitem.setBackground(Qt.lightGray)
            else:
                dateitem.setBackground(QColor(255, 204, 204))
                dateitem.setForeground(QColor(156, 0, 6))
 
            self.dailyviewtable.setItem(rowindex, 0, dateitem)
 
            for colindex, shippingdate in enumerate(shippingdates, 1):
                quantity = calloffs.get(shippingdate, 0)
 
                if quantity > 0:
                    qtyitem = QTableWidgetItem(f"{int(quantity):,}")
                    qtyitem.setBackground(Qt.cyan)
                else:
                    qtyitem = QTableWidgetItem("")
 
                qtyitem.setFlags(qtyitem.flags() & ~Qt.ItemIsEditable)
                self.dailyviewtable.setItem(rowindex, colindex, qtyitem)
 
    def displaydailydeltatable(self, waterfalldata, shippingdates):
        if not waterfalldata or not shippingdates:
            return
 
        headers = ['Call off Date'] + [d.strftime('%m/%d') for d in shippingdates]
 
        self.dailydeltatable.setRowCount(len(waterfalldata))
        self.dailydeltatable.setColumnCount(len(headers))
        self.dailydeltatable.setHorizontalHeaderLabels(headers)
        self.dailydeltatable.setColumnWidth(0, 120)
        for col in range(1, len(headers)):
            self.dailydeltatable.setColumnWidth(col, 80)
 
        for rowindex, rowdata in enumerate(waterfalldata):
            dateitem = QTableWidgetItem(rowdata['archive_date'].strftime('%m/%d'))
            dateitem.setFlags(dateitem.flags() & ~Qt.ItemIsEditable)
            dateitem.setBackground(Qt.lightGray)
            self.dailydeltatable.setItem(rowindex, 0, dateitem)
 
            if rowindex == 0:
                for colindex in range(1, len(headers)):
                    item = QTableWidgetItem("")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.dailydeltatable.setItem(rowindex, colindex, item)
                continue
 
            current = rowdata['call_offs']
            previous = waterfalldata[rowindex - 1]['call_offs']
 
            for colindex, shippingdate in enumerate(shippingdates, 1):
                delta = current.get(shippingdate, 0) - previous.get(shippingdate, 0)
 
                if delta != 0:
                    qtyitem = QTableWidgetItem(f"{int(delta):+,}")
                    if delta > 0:
                        qtyitem.setBackground(QColor(204, 255, 204))
                        qtyitem.setForeground(QColor(0, 156, 6))
                    else:
                        qtyitem.setBackground(QColor(255, 204, 204))
                        qtyitem.setForeground(QColor(156, 0, 6))
                else:
                    qtyitem = QTableWidgetItem("")
 
                qtyitem.setFlags(qtyitem.flags() & ~Qt.ItemIsEditable)
                self.dailydeltatable.setItem(rowindex, colindex, qtyitem)
 
    def displayweeklyviewtable(self, waterfalldata, shippingdates):
        if not waterfalldata or not shippingdates:
            return
 
        weeks = {}
        for d in shippingdates:
            iso = d.isocalendar()
            key = (iso[0], iso[1])
            weeks.setdefault(key, []).append(d)
 
        sortedweeks = sorted(weeks.keys())
        headers = ['Call off Date']
 
        for (yr, wk) in sortedweeks:
            weekdates = weeks[(yr, wk)]
            weekstart = min(weekdates)
            headers.append(f"Wk {wk}\n{weekstart.strftime('%m/%d')}")
 
        self.weeklyviewtable.setRowCount(len(waterfalldata))
        self.weeklyviewtable.setColumnCount(len(headers))
        self.weeklyviewtable.setHorizontalHeaderLabels(headers)
        self.weeklyviewtable.setColumnWidth(0, 120)
 
        for col in range(1, len(headers)):
            self.weeklyviewtable.setColumnWidth(col, 90)
 
        for rowindex, rowdata in enumerate(waterfalldata):
            archivedate = rowdata['archive_date']
            calloffs = rowdata['call_offs']
 
            dateitem = QTableWidgetItem(archivedate.strftime('%m/%d'))
            dateitem.setFlags(dateitem.flags() & ~Qt.ItemIsEditable)
            dateitem.setBackground(Qt.lightGray)
            self.weeklyviewtable.setItem(rowindex, 0, dateitem)
 
            for colindex, (yr, wk) in enumerate(sortedweeks, 1):
                weektotal = sum(calloffs.get(d, 0) for d in weeks[(yr, wk)])
 
                if weektotal > 0:
                    qtyitem = QTableWidgetItem(f"{int(weektotal):,}")
                    qtyitem.setBackground(Qt.cyan)
                else:
                    qtyitem = QTableWidgetItem("")
 
                qtyitem.setFlags(qtyitem.flags() & ~Qt.ItemIsEditable)
                self.weeklyviewtable.setItem(rowindex, colindex, qtyitem)
 
    def displayweeklydeltatable(self, waterfalldata, shippingdates):
        if not waterfalldata or not shippingdates:
            return
 
        weeks = {}
        for d in shippingdates:
            iso = d.isocalendar()
            key = (iso[0], iso[1])
            weeks.setdefault(key, []).append(d)
        sorted_weeks = sorted(weeks.keys())
 
        headers = ['Call off Date']
        for (yr, wk) in sorted_weeks:
            week_start = min(weeks[(yr, wk)])
            headers.append(f"Wk {wk}\n{week_start.strftime('%m/%d')}")
 
        self.weeklydeltatable.setRowCount(len(waterfalldata))
        self.weeklydeltatable.setColumnCount(len(headers))
        self.weeklydeltatable.setHorizontalHeaderLabels(headers)
        self.weeklydeltatable.setColumnWidth(0, 120)
        for col in range(1, len(headers)):
            self.weeklydeltatable.setColumnWidth(col, 90)
 
        for rowindex, rowdata in enumerate(waterfalldata):
            dateitem = QTableWidgetItem(rowdata['archive_date'].strftime('%m/%d'))
            dateitem.setFlags(dateitem.flags() & ~Qt.ItemIsEditable)
            dateitem.setBackground(Qt.lightGray)
            self.weeklydeltatable.setItem(rowindex, 0, dateitem)
 
            if rowindex == 0:
                for colindex in range(1, len(headers)):
                    item = QTableWidgetItem("")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.weeklydeltatable.setItem(rowindex, colindex, item)
                continue
 
            current = waterfalldata[rowindex]['call_offs']
            previous = waterfalldata[rowindex - 1]['call_offs']
 
            for colindex, (yr, wk) in enumerate(sorted_weeks, 1):
                current_total = sum(current.get(d, 0) for d in weeks[(yr, wk)])
                prev_total = sum(previous.get(d, 0) for d in weeks[(yr, wk)])
                delta = current_total - prev_total
 
                if delta != 0:
                    qtyitem = QTableWidgetItem(f"{int(delta):+,}")
                    if delta > 0:
                        qtyitem.setBackground(QColor(204, 255, 204))
                        qtyitem.setForeground(QColor(0, 156, 5))
                    else:
                        qtyitem.setBackground(QColor(255, 204, 204))
                        qtyitem.setForeground(QColor(156, 0, 6))
                else:
                    qtyitem = QTableWidgetItem("")
 
                qtyitem.setFlags(qtyitem.flags() & ~Qt.ItemIsEditable)
                self.weeklydeltatable.setItem(rowindex, colindex, qtyitem)
 
    def displaypartinfo(self, partinfo):
        for fieldname, value in partinfo.items():
            if fieldname in self.partinfolabels:
                if fieldname in ('Initial Stock', 'Unit Load Qty', 'Multi Unit Load', 'Safety Stock'):
                    display_value = f"{int(value):,}" if value else "0"
                elif fieldname == 'Piece Price':
                    display_value = f"${float(value):.2f}" if value else "$0.00"
                elif fieldname == 'Safety Days':
                    display_value = f"{float(value):.2f}" if value else "0.00"
                else:
                    display_value = str(value) if value else "--"
 
                self.partinfolabels[fieldname].setText(display_value)
 
    def clearpartsearch(self):
        self.partnumbersearch.clear()
        for label in self.partinfolabels.values():
            label.setText("--")
        self.transactiontable.setRowCount(0)
        if hasattr(self, 'partnotestextbox'):
            self.partnotestextbox.blockSignals(True)
            self.partnotestextbox.clear()
            self.partnotestextbox.blockSignals(False)
        self._currentpartnumber = None

    def _createpartnotessection(self):
        widget = QWidget()
        widget.setMaximumHeight(self._sz(200))
        layout = QVBoxLayout()
        layout.setContentsMargins(4, 0, 0, 0)

        notetitle = QLabel("Part Notes")
        notetitle.setFont(QFont("Arial", max(9, self._sz(14)), QFont.Bold))
        layout.addWidget(notetitle)

        self.partnotestextbox = QTextEdit()
        self.partnotestextbox.setPlaceholderText("Enter notes for this part...")
        self.partnotestextbox.setStyleSheet(
            "QTextEdit {border: 1px solid #ccc; border-radius: 4px; padding: 4px; background: #fffff5;}"
        )
        layout.addWidget(self.partnotestextbox)

        self._notessavetimer = QTimer()
        self._notessavetimer.setSingleShot(True)
        self._notessavetimer.setInterval(600)
        self._notessavetimer.timeout.connect(self._savecurrentpartnote)
        self.partnotestextbox.textChanged.connect(self._notessavetimer.start)

        widget.setLayout(layout)
        return widget

    def _partnotesfile(self):
        return getsharednetworkpath() / "partnotes.json"

    def _loadpartnotes(self):
        f = self._partnotesfile()
        if f.exists():
            try:
                with open(f, 'r') as fp:
                    return json.load(fp)
            except Exception as e:
                print(f"Error loading part notes: {e}")
        return {}

    def _savepartnotes(self, notes: dict):
        f = self._partnotesfile()
        try:
            f.parent.mkdir(parents=True, exist_ok=True)
            with open(f, 'w') as fp:
                json.dump(notes, fp, indent=2)
        except Exception as e:
            print(f"Error saving part notes: {e}")

    def _loadnoteforpart(self, partnumber: str):
        notes = self._loadpartnotes()
        self.partnotestextbox.blockSignals(True)
        self.partnotestextbox.setPlainText(notes.get(partnumber, ''))
        self.partnotestextbox.blockSignals(False)

    def _savecurrentpartnote(self):
        if not hasattr(self, '_currentpartnumber') or not self._currentpartnumber:
            return
        notes = self._loadpartnotes()
        text = self.partnotestextbox.toPlainText().strip()
        if text:
            notes[self._currentpartnumber] = text
        else:
            notes.pop(self._currentpartnumber, None)
        self._savepartnotes(notes)

    def clearcalloffanalysis(self):
        self.calloffpartsearch.clear()
        self.dailyviewtable.setRowCount(0)
        self.dailydeltatable.setRowCount(0)
        self.weeklyviewtable.setRowCount(0)
        self.weeklydeltatable.setRowCount(0)
 
    def generatecoverageanalysis(self):
        self._comments_cache = None
        try:
            success, message, datadict = self.coverageengine.loadrequireddata()
            if not success:
                QMessageBox.warning(self, "Missing Data", message)
                return
 
            coveragedf = self.coverageengine.buildcoverageanalysis(datadict, target_consumption_days=100)
            # buildcoverageanalysis already calls addcoveragecomments internally;
            # no second mapping needed here

            if coveragedf.empty:
                QMessageBox.information(self, "No Data", "No parts with consumption found.")
                return
 
            self.originalcoveragedf = coveragedf.copy()
            self.currentcoveragedf = coveragedf
            self.populatefilters(coveragedf)
            self.displaycoveragetable(coveragedf)
 
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Analysis failed: {str(e)}")
 
    def generateldjiscoverage(self):
        try:
            success, message, datadict = self.coverageengine.loadrequireddata()
            if not success:
                QMessageBox.warning(self, "Missing Data", message)
                return
 
            ok, msg, ldjisdf = self.ldjiscoverageengine.buildldjiscoveragedata(datadict)
            if not ok:
                QMessageBox.warning(self, "No LDJIS Data", msg)
                return
 
            self.ldjisdf = ldjisdf
            self.ldjisworkingdays = self.ldjiscoverageengine.generateworkingdays(
                datetime.now().date(), 20
            )
            self.displayldjistable()
 
        except Exception as e:
            QMessageBox.critical(self, "Error", f"LDJIS analysis failed: {str(e)}")
 
    def displayldjistable(self):
        if not hasattr(self, 'ldjisdf') or self.ldjisdf.empty:
            return
 
        try:
            self.ldjistable.itemChanged.disconnect()
        except Exception:
            pass
 
        nmodels = len(self.ldjismodelnames)
        nparts = len(self.ldjisdf)
        nfixed = self.ldjiscoverageengine.N_FIXED
        fixedcols = self.ldjiscoverageengine.FIXED_COLS
        datecols = [d.strftime('%d-%b') for d in self.ldjisworkingdays]
        allheaders = fixedcols + datecols
 
        self.ldjistable.horizontalHeader().setVisible(False)
 
        self.ldjistable.setRowCount(1 + nmodels + nparts)
        self.ldjistable.setColumnCount(len(allheaders))
        vlabels = [''] + [str(i) for i in range(1, nmodels + nparts + 1)]
        self.ldjistable.setVerticalHeaderLabels(vlabels)
 
        for col, width in enumerate([230, 80, 130, 130, 130]):
            self.ldjistable.setColumnWidth(col, width)
        for col in range(nfixed, len(allheaders)):
            self.ldjistable.setColumnWidth(col, 90)
 
        headerfont = QFont("Arial", 9, QFont.Bold)
 
        for col, label in enumerate(fixedcols):
            item = QTableWidgetItem(label)
            item.setFont(headerfont)
            item.setBackground(Qt.lightGray)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.ldjistable.setItem(0, col, item)
            self.ldjistable.setSpan(0, col, nmodels + 1, 1)
 
        for col, label in enumerate(datecols, nfixed):
            item = QTableWidgetItem(label)
            item.setFont(headerfont)
            item.setBackground(Qt.lightGray)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.ldjistable.setItem(0, col, item)
 
        self.ldjistable.setRowHeight(0, 35)
 
        italicfont = QFont("Arial", 9)
        italicfont.setItalic(True)
 
        for modelrow, modelname in enumerate(self.ldjismodelnames):
            tablerow = 1 + modelrow
 
            nameitem = QTableWidgetItem(f"{modelname} volume ->")
            nameitem.setFont(italicfont)
            nameitem.setForeground(Qt.darkGray)
            nameitem.setFlags(nameitem.flags() | Qt.ItemIsEditable)
            self.ldjistable.setItem(tablerow, nfixed, nameitem)
 
            for col in range(nfixed + 1, len(allheaders)):
                volitem = QTableWidgetItem("0")
                volitem.setFlags(volitem.flags() | Qt.ItemIsEditable)
                self.ldjistable.setItem(tablerow, col, volitem)
 
        for partrow, (_, row) in enumerate(self.ldjisdf.iterrows()):
            tablerow = 1 + nmodels + partrow
            fixedvals = [
                str(row['Part Number']),
                str(row['Part Description']),
                str(row['SHP Code']),
                f"{row['Moves away from COL']:,.0f}",
                f"{int(row['Last Covered Mix']):,}",
                f"{int(row['Starting COL Mix']):,}",
            ]
            for col, val in enumerate(fixedvals):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.ldjistable.setItem(tablerow, col, item)
 
            for col in range(nfixed, len(allheaders)):
                item = QTableWidgetItem("")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.ldjistable.setItem(tablerow, col, item)
 
        self.recalculateldjiscoverage()
        self.ldjistable.itemChanged.connect(self.onldjisvolchanged)
 
    def recalculateldjiscoverage(self):
        if not hasattr(self, 'ldjisdf') or self.ldjisdf.empty:
            return
 
        try:
            self.ldjistable.itemChanged.disconnect()
        except Exception:
            pass
 
        try:
            nmodels = len(self.ldjismodelnames)
            nfixed = self.ldjiscoverageengine.N_FIXED
            ncols = self.ldjistable.columnCount()
            ndays = ncols - nfixed
 
            dayvolumes = np.zeros(ndays)
            for daycol in range(ndays):
                tablecol = nfixed + daycol
                for modelrow in range(1, 1 + nmodels):
                    item = self.ldjistable.item(modelrow, tablecol)
                    if item:
                        try:
                            dayvolumes[daycol] += float(item.text().replace(',', '').strip() or 0)
                        except ValueError:
                            pass
 
            moves_away = self.ldjisdf['Moves away from COL'].astype(float).values
            starting_mix = self.ldjisdf['Starting COL Mix'].astype(float).values
            n_parts = len(moves_away)
 
            coverage = np.empty((n_parts, ndays))
            coverage[:, 0] = starting_mix - moves_away - dayvolumes[0]
            for d in range(1, ndays):
                coverage[:, d] = coverage[:, d - 1] - dayvolumes[d]
 
            for partrow in range(n_parts):
                tablerow = 1 + nmodels + partrow
                for daycol in range(ndays):
                    val = int(coverage[partrow, daycol])
                    item = QTableWidgetItem(f"{val:,}")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if val <= 0:
                        item.setBackground(Qt.red)
                        item.setForeground(Qt.white)
                    else:
                        item.setBackground(Qt.white)
                        item.setForeground(Qt.black)
                    self.ldjistable.setItem(tablerow, nfixed + daycol, item)
 
        finally:
            self.ldjistable.itemChanged.connect(self.onldjisvolchanged)
 
    def onldjisvolchanged(self, item):
        nmodels = len(self.ldjismodelnames)
        nfixed = self.ldjiscoverageengine.N_FIXED
        if 1 <= item.row() <= nmodels and item.column() >= nfixed:
            self.recalculateldjiscoverage()
 
    def displaycoveragetable(self, coveragedf: pd.DataFrame):
        if coveragedf.empty:
            return
 
        try:
            self.coveragetable.itemChanged.disconnect()
        except Exception:
            pass
 
        cols = coveragedf.columns.tolist()
        n_rows, n_cols = len(coveragedf), len(cols)
        str_cols = {cols.index(c) for c in ('Part Number', 'MFG Code', 'SHP Code') if c in cols}
 
        if not hasattr(self, '_last_column_count') or self._last_column_count != n_cols:
            self.coveragetable.setColumnCount(n_cols)
            self.coveragetable.setHorizontalHeaderLabels(cols)
            self._last_column_count = n_cols
            self.comments_col = cols.index('Comments') if 'Comments' in cols else None
 
        self.coveragetable.setSortingEnabled(False)
        self.coveragetable.setUpdatesEnabled(False)
 
        try:
            self.coveragetable.setRowCount(n_rows)
 
            data = coveragedf.values
 
            daily_start = None
            for i, col in enumerate(cols):
                try:
                    datetime.strptime(col, '%m/%d')
                    daily_start = i
                    break
                except ValueError:
                    pass
 
            unit_loads = (
                coveragedf['Unit Load Qty'].fillna(1).clip(lower=1).values
                if 'Unit Load Qty' in cols
                else None
            )
 
            for row in range(n_rows):
                ulq = unit_loads[row] if unit_loads is not None else 1
 
                for col in range(n_cols):
                    value = data[row, col]
                    isnumeric = isinstance(value, (int, float)) and value == value
 
                    if isnumeric and col not in str_cols:
                        display_value = f"{value:,.0f}" if abs(value) >= 1 else f"{value:.2f}"
                    elif isinstance(value, float) and value != value:
                        display_value = ""
                    else:
                        display_value = str(value) if value is not None else ""
 
                    if cols[col] == 'Day Alert':
                        try:
                            sort_value = int(value)
                        except (ValueError, TypeError):
                            text = str(display_value).strip()
                            if text.lower() == 'covered':
                                sort_value = 999
                            else:
                                m = re.search(r'(\d+)', text)
                                sort_value = int(m.group(1)) if m else 999999
                        item = NumericSortTableWidgetItem(display_value, sort_value)
                    else:
                        item = QTableWidgetItem(display_value)
                    item.setFlags(
                        (item.flags() | Qt.ItemIsEditable)
                        if col == self.comments_col
                        else (item.flags() & ~Qt.ItemIsEditable)
                    )
 
                    if daily_start is not None and col >= daily_start and isnumeric:
                        if value <= 0:
                            item.setBackground(QColor(255, 204, 204))
                            item.setForeground(QColor(156, 0, 6))
                        elif value < ulq:
                            item.setBackground(QColor(255, 255, 204))
                            item.setForeground(QColor(156, 156, 6))
 
                    self.coveragetable.setItem(row, col, item)
 
        finally:
            self.coveragetable.setUpdatesEnabled(True)
            self.coveragetable.resizeColumnsToContents()
            if self.comments_col is not None:
                self.coveragetable.setColumnWidth(self.comments_col, 200)
                if 'Comments' in coveragedf.columns:
                    comments_vals = coveragedf['Comments'].values
                    for row in range(n_rows):
                        if str(comments_vals[row]).strip():
                            self.coveragetable.resizeRowToContents(row)
            self.coveragetable.setSortingEnabled(True)
            self.coveragetable.itemChanged.connect(self.oncommentchanged)
            self._reapplycoveragehiddencolumns() 
            self._applyfrozencolumns()
 
    def oncommentchanged(self, item):
        try:
            col = item.column()
            if 'Comments' not in self.currentcoveragedf.columns:
                return
 
            commentscol = list(self.currentcoveragedf.columns).index('Comments')
            if col != commentscol:
                return
 
            partcol = list(self.currentcoveragedf.columns).index('Part Number')
            part_item = self.coveragetable.item(item.row(), partcol)
            if not part_item:
                return
            partno = re.sub(r'\.0$', '', part_item.text().strip())
            commenttext = item.text().strip()
 
            if self._comments_cache is None:
                self._comments_cache = self.coverageengine.loadcoveragecomments()
            if self._coverage_pending is None:
                self._coverage_pending = self.coverageengine.loadpendingcoveragecomments()

            if commenttext:
                self._comments_cache[partno] = commenttext
            else:
                self._comments_cache.pop(partno, None)

            # Track as pending (empty string = delete on upload) and save locally only
            self._coverage_pending[partno] = commenttext
            self.coverageengine.savecoveragecomments(self._coverage_pending)
            self.coveragetable.resizeRowToContents(item.row())
         
            if hasattr(self, 'originalcoveragedf') and 'Comments' in self.originalcoveragedf.columns:
                mask = self.originalcoveragedf['Part Number'].astype(str).str.replace(r'\.0$', '', regex=True) == partno
                self.originalcoveragedf.loc[mask, 'Comments'] = commenttext

        except Exception as e:
            print(f"Error saving comment: {e}")

    def _uploadcoveragecomments(self):
        if self._coverage_pending is None:
            self._coverage_pending = self.coverageengine.loadpendingcoveragecomments()
        if not self._coverage_pending:
            QMessageBox.information(self, "Upload Comments", "No pending comments to upload.")
            return
        success = self.coverageengine.uploadcoveragecomments(self._coverage_pending)
        if success:
            self._coverage_pending = {}
            QMessageBox.information(self, "Upload Comments", "Comments uploaded to the shared file successfully.")
        else:
            QMessageBox.warning(self, "Upload Failed",
                "Could not write to the shared file — it may be in use by another user.\n"
                "Your comments are saved locally. Please try again in a moment.")

    def _alertsfile(self):
        return getsharednetworkpath() / "alertsdata.json"

    def _alertslocalfile(self):
        return LOCALAPPDATA / "alertsdata_pending.json"

    def _loadalertsdata(self):
        # Load shared baseline
        shared = {}
        f = self._alertsfile()
        if f.exists():
            try:
                with open(f, 'r') as fp:
                    shared = json.load(fp)
            except Exception as e:
                print(f"Error loading alerts data: {e}")
        # Deep-merge local pending on top (empty string = field was deleted)
        lf = self._alertslocalfile()
        if lf.exists():
            try:
                with open(lf, 'r') as fp:
                    local = json.load(fp)
                for key, fields in local.items():
                    if key not in shared:
                        shared[key] = {}
                    for col, val in fields.items():
                        if val:
                            shared[key][col] = val
                        else:
                            shared[key].pop(col, None)
                    if not shared.get(key):
                        shared.pop(key, None)
            except Exception as e:
                print(f"Error loading local pending alerts data: {e}")
        return shared

    def _savealertsdata(self, pending: dict):
        # Save ONLY the user's pending changes to local file, not the shared file
        lf = self._alertslocalfile()
        try:
            lf.parent.mkdir(parents=True, exist_ok=True)
            with open(lf, 'w') as fp:
                json.dump(pending, fp, indent=2)
        except Exception as e:
            print(f"Error saving local pending alerts data: {e}")

    def _uploadalertsdata(self):
        if self._alerts_pending is None:
            self._alerts_pending = {}
        if not self._alerts_pending:
            QMessageBox.information(self, "Upload Field Data", "No pending field data to upload.")
            return
        f = self._alertsfile()
        try:
            shared = {}
            if f.exists():
                with open(f, 'r') as fp:
                    shared = json.load(fp)
            # Deep merge: apply only user's changes onto current shared state
            for key, fields in self._alerts_pending.items():
                if key not in shared:
                    shared[key] = {}
                for col, val in fields.items():
                    if val:
                        shared[key][col] = val
                    else:
                        shared[key].pop(col, None)
                if not shared.get(key):
                    shared.pop(key, None)
            f.parent.mkdir(parents=True, exist_ok=True)
            with open(f, 'w') as fp:
                json.dump(shared, fp, indent=2)
            # Clear local pending
            self._alerts_pending = {}
            self._alertslocalfile().write_text('{}')
            QMessageBox.information(self, "Upload Field Data", "Field data uploaded to the shared file successfully.")
        except Exception as e:
            print(f"Error uploading alerts data: {e}")
            QMessageBox.warning(self, "Upload Failed",
                "Could not write to the shared file — it may be in use by another user.\n"
                "Your field data is saved locally. Please try again in a moment.")

    def _alerthighlightsfile(self):
        return getsharednetworkpath() / "alert_highlights.json"

    def _loadalerthighlights(self) -> set:
        f = self._alerthighlightsfile()
        if f.exists():
            try:
                with open(f, 'r') as fp:
                    return set(json.load(fp))
            except Exception as e:
                print(f"Error loading alert highlights: {e}")
        return set()

    def _savealerthighlights(self, highlights: set):
        f = self._alerthighlightsfile()
        try:
            f.parent.mkdir(parents=True, exist_ok=True)
            with open(f, 'w') as fp:
                json.dump(list(highlights), fp)
        except Exception as e:
            print(f"Error saving alert highlights: {e}")

    def _onalerts_contextmenu(self, pos):
        item = self.alertstable.itemAt(pos)
        if item is None:
            return
        row = item.row()
        cols = [self.alertstable.horizontalHeaderItem(c).text() if self.alertstable.horizontalHeaderItem(c) else ''
                for c in range(self.alertstable.columnCount())]
        part_idx = cols.index('Part') if 'Part' in cols else -1
        if part_idx < 0:
            return
        part_item = self.alertstable.item(row, part_idx)
        if not part_item:
            return
        part = part_item.text()
        if self._alert_highlights is None:
            self._alert_highlights = self._loadalerthighlights()
        menu = QMenu(self)
        if part in self._alert_highlights:
            action = menu.addAction("Remove Highlight")
            action.triggered.connect(lambda: self._togglealertrowhighlight(part, False))
        else:
            action = menu.addAction("Highlight Row")
            action.triggered.connect(lambda: self._togglealertrowhighlight(part, True))
        menu.exec_(self.alertstable.viewport().mapToGlobal(pos))

    def _togglealertrowhighlight(self, part, highlight):
        if self._alert_highlights is None:
            self._alert_highlights = self._loadalerthighlights()
        if highlight:
            self._alert_highlights.add(part)
        else:
            self._alert_highlights.discard(part)
        self._savealerthighlights(self._alert_highlights)
        self._applyalerttablehighlights()

    def _applyalerttablehighlights(self):
        cols = [self.alertstable.horizontalHeaderItem(c).text() if self.alertstable.horizontalHeaderItem(c) else ''
                for c in range(self.alertstable.columnCount())]
        part_idx = cols.index('Part') if 'Part' in cols else -1
        alerts_col_idx = cols.index('Alerts') if 'Alerts' in cols else -1
        if part_idx < 0 or alerts_col_idx < 0:
            return
        highlights = self._alert_highlights or set()
        is_admin = self.userdata.get('username', '') in ADMINUSERS
        try:
            self.alertstable.itemChanged.disconnect()
        except Exception:
            pass
        for row in range(self.alertstable.rowCount()):
            part_item = self.alertstable.item(row, part_idx)
            part = part_item.text() if part_item else ''
            cell = self.alertstable.item(row, alerts_col_idx)
            if cell is None:
                continue
            if part in highlights:
                cell.setBackground(QColor(255, 235, 59))
            elif is_admin:
                cell.setBackground(QColor(230, 245, 255))
            else:
                cell.setBackground(QColor(255, 255, 255))
        self.alertstable.itemChanged.connect(self._onalertchanged)

    def _onalertchanged(self, item):
        try:
            cols = [self.alertstable.horizontalHeaderItem(c).text() if self.alertstable.horizontalHeaderItem(c) else ''
                    for c in range(self.alertstable.columnCount())]
            if item.column() >= len(cols):
                return
            col_name = cols[item.column()]

            part_idx = cols.index('Part') if 'Part' in cols else -1
            alert_idx = cols.index('Alerts') if 'Alerts' in cols else -1
            if part_idx < 0 or alert_idx < 0:
                return
            part_item = self.alertstable.item(item.row(), part_idx)
            alert_item = self.alertstable.item(item.row(), alert_idx)
            if not part_item or not alert_item:
                return

            # Admin editing the Alerts (day) value — update the df but don't persist
            if col_name == 'Alerts' and self.userdata.get('username', '') in ADMINUSERS:
                if hasattr(self, 'originalalertsdf') and 'Part' in self.originalalertsdf.columns:
                    mask = self.originalalertsdf['Part'].astype(str) == part_item.text()
                    self.originalalertsdf.loc[mask, 'Alerts'] = item.text().strip()
                return

            if col_name not in self._ALERT_EDITABLE_COLS:
                return

            key = re.sub(r'\.0$', '', part_item.text().strip())

            if self._alerts_cache is None:
                self._alerts_cache = self._loadalertsdata()
            if self._alerts_pending is None:
                lf = self._alertslocalfile()
                self._alerts_pending = json.loads(lf.read_text()) if lf.exists() else {}

            if key not in self._alerts_cache:
                self._alerts_cache[key] = {}
            val = item.text().strip()
            if val:
                self._alerts_cache[key][col_name] = val
            else:
                self._alerts_cache[key].pop(col_name, None)
                if not self._alerts_cache[key]:
                    del self._alerts_cache[key]

            # Track as pending (empty string = delete on upload) and save locally only
            if key not in self._alerts_pending:
                self._alerts_pending[key] = {}
            self._alerts_pending[key][col_name] = val  # empty string marks deletion
            self._savealertsdata(self._alerts_pending)
            self.alertstable.resizeRowToContents(item.row())
         
            if hasattr(self, 'originalalertsdf') and col_name in self.originalalertsdf.columns:
                if 'Part' in self.originalalertsdf.columns:
                    part_norm = re.sub(r'\.0$', '', part_item.text().strip())
                    orig_parts = self.originalalertsdf['Part'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                    mask = orig_parts == part_norm
                    self.originalalertsdf.loc[mask, col_name] = item.text().strip()

        except Exception as e:
            print(f"Error saving alert change: {e}")

    def _piwdfile(self):
        return getsharednetworkpath() / "piwddata.json"

    def _loadpiwddata(self):
        f = self._piwdfile()
        if f.exists():
            try:
                with open(f, 'r') as fp:
                    return json.load(fp)
            except Exception as e:
                print(f"Error loading PIWD data: {e}")
        return {}

    def _savepiwddata(self, data: dict):
        f = self._piwdfile()
        try:
            f.parent.mkdir(parents=True, exist_ok=True)
            with open(f, 'w') as fp:
                json.dump(data, fp, indent=2)
        except Exception as e:
            print(f"Error saving PIWD data: {e}")

    def _onpiwdchanged(self, item):
        try:
            cols = [self.piwdtable.horizontalHeaderItem(c).text() if self.piwdtable.horizontalHeaderItem(c) else ''
                    for c in range(self.piwdtable.columnCount())]
            if item.column() >= len(cols):
                return
            col_name = cols[item.column()]
            if col_name not in self._PIWD_EDITABLE_COLS:
                return

            part_idx = cols.index('Part') if 'Part' in cols else -1
            if part_idx < 0:
                return
            part_item = self.piwdtable.item(item.row(), part_idx)
            if not part_item:
                return
            key = part_item.text()

            if self._piwd_cache is None:
                self._piwd_cache = self._loadpiwddata()

            if key not in self._piwd_cache:
                self._piwd_cache[key] = {}
            val = item.text().strip()
            if val:
                self._piwd_cache[key][col_name] = val
            else:
                self._piwd_cache[key].pop(col_name, None)
                if not self._piwd_cache[key]:
                    del self._piwd_cache[key]

            self._savepiwddata(self._piwd_cache)
            self.piwdtable.resizeRowToContents(item.row())
        except Exception as e:
            print(f"Error saving PIWD change: {e}")

    def displaytransactiontable(self, transactiondata):
        if not transactiondata:
            return
 
        self.transactiontable.setRowCount(len(transactiondata))
 
        for row, transaction in enumerate(transactiondata):
            dateitem = QTableWidgetItem(transaction['Date'])
            dateitem.setFlags(dateitem.flags() & ~Qt.ItemIsEditable)
            self.transactiontable.setItem(row, 0, dateitem)
 
            typeitem = QTableWidgetItem(transaction['Transaction Type'])
            typeitem.setFlags(typeitem.flags() & ~Qt.ItemIsEditable)
 
            if transaction['Transaction Type'] == 'Stock':
                typeitem.setBackground(Qt.lightGray)
            elif transaction['Transaction Type'] == 'Req':
                typeitem.setBackground(QColor(255, 204, 204))
                typeitem.setForeground(QColor(156, 0, 6))
            elif transaction['Transaction Type'] == 'ASN':
                typeitem.setBackground(QColor(204, 255, 204))
                typeitem.setForeground(QColor(0, 156, 6))
 
            self.transactiontable.setItem(row, 1, typeitem)
 
            receiptitem = QTableWidgetItem(transaction['Receipt/Reqmt'])
            receiptitem.setFlags(receiptitem.flags() & ~Qt.ItemIsEditable)
            self.transactiontable.setItem(row, 2, receiptitem)
 
            qtyitem = QTableWidgetItem(f"{transaction['Available QTY']:,}")
            qtyitem.setFlags(qtyitem.flags() & ~Qt.ItemIsEditable)
 
            if transaction['Available QTY'] <= 0:
                qtyitem.setBackground(QColor(255, 204, 204))
                qtyitem.setForeground(QColor(156, 0, 6))
            elif transaction['Available QTY'] < 100:
                qtyitem.setBackground(QColor(255, 255, 204))
                qtyitem.setForeground(QColor(156, 156, 6))
 
            self.transactiontable.setItem(row, 3, qtyitem)
 
            asnitem = QTableWidgetItem(transaction['ASN'])
            asnitem.setFlags(asnitem.flags() & ~Qt.ItemIsEditable)
            self.transactiontable.setItem(row, 4, asnitem)

    def _export_table_to_excel(self, table, filename, sheet_name='Sheet1'):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        n_cols = table.columnCount()
        n_rows = table.rowCount()
        for col in range(n_cols):
            header = table.horizontalHeaderItem(col)
            ws.cell(row=1, column=col + 1, value=header.text() if header else '')
        for row in range(n_rows):
            for col in range(n_cols):
                cell_item = table.item(row, col)
                value = cell_item.text() if cell_item else ''
                ws_cell = ws.cell(row=row + 2, column=col + 1, value=value)
                if cell_item:
                    brush = cell_item.background()
                    if brush.style() != 0:  # 0 = Qt.NoBrush
                        c = brush.color()
                        r, g, b = c.red(), c.green(), c.blue()
                        if not (r == 255 and g == 255 and b == 255):
                            hex_color = f'{r:02X}{g:02X}{b:02X}'
                            ws_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                    fg = cell_item.foreground().color()
                    fr, fg2, fb = fg.red(), fg.green(), fg.blue()
                    if not (fr == 0 and fg2 == 0 and fb == 0):
                        ws_cell.font = Font(color=f'{fr:02X}{fg2:02X}{fb:02X}')
        wb.save(filename)

    def exporttransactiontable(self):
        if not hasattr(self, 'transactiontable') or self.transactiontable.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "Generate transaction breakdown first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Transaction Breakdown",
            f"Transaction_Breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx);;CSV files (*.csv)"
        )
        if not filename:
            return
        try:
            if filename.endswith('.csv'):
                cols = [self.transactiontable.horizontalHeaderItem(c).text() if self.transactiontable.horizontalHeaderItem(c) else f'Col{c}' for c in range(self.transactiontable.columnCount())]
                rows = [[self.transactiontable.item(r, c).text() if self.transactiontable.item(r, c) else '' for c in range(self.transactiontable.columnCount())] for r in range(self.transactiontable.rowCount())]
                pd.DataFrame(rows, columns=cols).to_csv(filename, index=False)
            else:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                self._export_table_to_excel(self.transactiontable, filename, 'Transaction Breakdown')
            QMessageBox.information(self, "Export Complete", f"Exported to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))
 
    _ALERT_COL_MAP = {
        'SCC_NAME':            'SCC',
        'ALERT_DETAILS':       'Alerts',
        'PART':                'Part',
        'PART_DESCRIPTION':    'Part Description',
        'CURRENT_INVENTORY':   'Inv',
        'ON_YARD_INVENTORY':   'Yard',
        'CURRENT_REQUIREMENT': 'Req',
        'SUPPLIER_COUNTRY':    'Country',
        'SUPPLIER_NAME':       'Supplier',
        }
 
    _ALERT_EDITABLE_COLS = [
        'M',
        'Reason/Cause',
        '#Cars Short',
        'Impact Date',
        'ETA',
        'QTY',
        'TO/Container#/Air Freight/Delivery Time',
        'ASN Y/N',
        'Unloading Shop A/C/LOCD',
        ]
 
    def generatealertsbreakdown(self):
        try:
            alertdf = self.import_manager.loaddata("alert_report")
            if alertdf.empty:
                QMessageBox.warning(self, "No Data", "No Alert Report found. Please import it first.")
                return
            
            if 'ALERT_TYPE' in alertdf.columns:
                shortage_df = alertdf[alertdf['ALERT_TYPE'] == 'Shortage alert']
            else:
                shortage_df = alertdf

            if 'ALERT_DETAILS' in alertdf.columns:
                today = datetime.now().date()

                day_mask = shortage_df['ALERT_DETAILS'].str.contains(r'Day [1-4]\b', case=False, na=False)
                normal_df = shortage_df[day_mask].copy()
                normal_df['ALERT_DETAILS'] = normal_df['ALERT_DETAILS'].str.extract(r'(Day [1-4])\b', expand=False)

                piwed_candidates = pd.DataFrame()
                piwed_mask = alertdf['ALERT_DETAILS'].str.contains(r'PIWED below zero using GC ETA', case=False, na=False) | \
                             alertdf['ALERT_DETAILS'].str.contains(r'PIWD below zero', case=False, na=False) | \
                            alertdf['ALERT_DETAILS'].str.contains(r'PIWDCS below zero', case=False, na=False)
                if piwed_mask.any() and 'COMMENTS' in alertdf.columns:
                    piwed_df = alertdf[piwed_mask].copy()

                    def _piwed_day_label(comments):
                        if not isinstance(comments, str):
                            return None
                        m = re.search(r'Prod\.day=(\d{4}-\d{2}-\d{2})', comments, re.IGNORECASE)
                        if not m:
                            return None
                        try:
                            diff = (datetime.strptime(m.group(1), '%Y-%m-%d').date() - today + timedelta(days=1)).days
                            return f'Day {diff}' if 1 <= diff <= 4 else None
                        except ValueError:
                            return None

                    piwed_df['ALERT_DETAILS'] = piwed_df['COMMENTS'].apply(_piwed_day_label)
                    piwed_candidates = piwed_df[piwed_df['ALERT_DETAILS'].notna()].copy()

                alertdf = pd.concat([normal_df, piwed_candidates], ignore_index=True)

                if 'PART' in alertdf.columns and not alertdf.empty:
                    # Normalise to clean string so float 12345.0 and str "12345"
                    # are treated as the same part by drop_duplicates.
                    alertdf['PART'] = (alertdf['PART'].astype(str).str.strip()
                                       .str.replace(r'\.0$', '', regex=True))
                    alertdf = alertdf.drop_duplicates(subset=['PART', 'ALERT_DETAILS'], keep='first')
                    alertdf['_day_num'] = alertdf['ALERT_DETAILS'].str.extract(r'Day (\d+)', expand=False).astype(float)
                    alertdf = alertdf.sort_values('_day_num').drop_duplicates(subset=['PART'], keep='first').drop(columns=['_day_num'])
            else:
                alertdf = shortage_df

            if alertdf.empty:
                QMessageBox.information(self, "No Data", "No Shortage alert rows for Day 1-4 found in the alert report")
                return
            
            available = {src: dst for src, dst in self._ALERT_COL_MAP.items() if src in alertdf.columns}
            displaydf = alertdf[list(available.keys())].rename(columns=available).copy()
 
            for col in self._ALERT_EDITABLE_COLS:
                displaydf[col] = ''

            saved = self._loadalertsdata()
            self._alerts_cache = saved  # always warm the cache after load
            if 'Part' in displaydf.columns:
                def _pkey(v):
                    return re.sub(r'\.0$', '', str(v).strip())
                current_keys = {_pkey(r['Part']) for _, r in displaydf.iterrows()}
                # prune keys whose part is no longer in any alert (normalised comparison)
                stale = [k for k in saved if _pkey(k) not in current_keys]
                for k in stale:
                    del saved[k]
                if stale:
                    self._savealertsdata(saved)
                # restore saved values regardless of whether stale keys existed
                for idx, row in displaydf.iterrows():
                    key = _pkey(row['Part'])
                    if key in saved:
                        for col, val in saved[key].items():
                            if col in displaydf.columns:
                                displaydf.at[idx, col] = val

            if 'Country' in displaydf.columns:
                displaydf['Region'] = displaydf['Country'].apply(
                    self.coverageengine.determineregion
                )
            else:
                displaydf['Region'] = ''

            try:
                partmatrix = self.import_manager.loaddata('part_matrix')
                if not partmatrix.empty and 'Part No' in partmatrix.columns:
                    v536_col = 'Type 110 (V536)'
                    p519_col = 'Type 100 (P519)'
                    def _classify_program(row):
                        has_v536 = v536_col in row.index and str(row[v536_col]).strip().upper() == 'X'
                        has_p519 = p519_col in row.index and str(row[p519_col]).strip().upper() == 'X'
                        if has_v536 and has_p519:
                            return 'Common'
                        elif has_v536:
                            return 'V536'
                        elif has_p519:
                            return 'P519'
                        return ''
                    pm = partmatrix.copy()
                    pm['_program'] = pm.apply(_classify_program, axis=1)
                    pm['_part_key'] = pm['Part No'].astype(str).str.strip().str.upper()
                    lookup = pm.set_index('_part_key')['_program'].to_dict()
                    if 'Part' in displaydf.columns:
                        displaydf['Program Supported'] = (
                            displaydf['Part'].astype(str).str.strip().str.upper().map(lookup).fillna('')
                        )
                else:
                    displaydf['Program Supported'] = ''
            except Exception:
                displaydf['Program Supported'] = ''

            preferred_order = [
                'M', 'SCC', 'Alerts', 'Part', 'Part Description', 'Inv', 'Yard', 'Req',
                'Country', 'Region', 'Program Supported', 'Supplier',
            ] + [c for c in self._ALERT_EDITABLE_COLS if c != 'M']
            ordered = [c for c in preferred_order if c in displaydf.columns]
            remaining = [c for c in displaydf.columns if c not in ordered]
            displaydf = displaydf[ordered + remaining]

            # Final dedup: normalise Part to clean string first so that float
            # 12345.0 and str "12345" are treated as the same value, then drop
            # any remaining duplicate part rows keeping the lowest alert day.
            if 'Part' in displaydf.columns and 'Alerts' in displaydf.columns:
                displaydf['Part'] = (displaydf['Part'].astype(str).str.strip()
                                     .str.replace(r'\.0$', '', regex=True))
                displaydf['_day_sort'] = displaydf['Alerts'].str.extract(r'Day (\d+)', expand=False).astype(float)
                displaydf = (displaydf
                             .sort_values('_day_sort', na_position='last')
                             .drop_duplicates(subset=['Part'], keep='first')
                             .drop(columns=['_day_sort'])
                             .reset_index(drop=True))

            self.originalalertsdf = displaydf.copy()
            self.populatealertfilters(displaydf)
            self.displayalertstable(displaydf)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Alerts analysis failed: {str(e)}")
            
    _PIWD_COL_MAP = {
        'SCC_NAME': 'SCC Name',
        'PART': 'Part',
        'PART_DESCRIPTION': 'Part Description',
        'CURRENT_INVENTORY': 'Inv',
        'ON_YARD_INVENTORY': 'Yard',
        'PORT_INVENTORY': 'Port',
        'CURRENT_REQUIREMENT': 'Req',
        'SUPPLIER_NAME': 'Supplier',
        'SUPPLY_SHP_COUNTRY': 'Country',
        }
    
    _PIWD_EDITABLE_COLS = [
        'Qty',
        'Impact with Material in Transit (No Gap)',
        'Comments'
        ]
            
    def generatepiwdreport(self):
        try:
            piwddf = self.import_manager.loaddata("alert_report")
            if piwddf.empty:
                QMessageBox.warning(self, "No Data", "No Alert Report found. Please import it first.")
                return
            
            if 'ALERT_DETAILS' in piwddf.columns:
                piwddf = piwddf[piwddf['ALERT_DETAILS'] == 'PIWED below zero using GC ETA']
            
            if piwddf.empty:
                QMessageBox.information(self, "No Data", "No PIWD alert rows in the alert report")
                return
            
            available = {src: dst for src, dst in self._PIWD_COL_MAP.items() if src in piwddf.columns}
            displaydf = piwddf[list(available.keys())].rename(columns=available).copy()
            
            for col in self._PIWD_EDITABLE_COLS:
                displaydf[col] = ''

            saved = self._loadpiwddata()
            if 'Part' in displaydf.columns:
                current_keys = {str(r['Part']) for _, r in displaydf.iterrows()}
                if saved:
                    stale = [k for k in saved if k not in current_keys]
                    for k in stale:
                        del saved[k]
                    if stale:
                        self._piwd_cache = saved
                        self._savepiwddata(saved)
                    for idx, row in displaydf.iterrows():
                        key = str(row['Part'])
                        if key in saved:
                            for col, val in saved[key].items():
                                if col in displaydf.columns:
                                    displaydf.at[idx, col] = val

            self.originalpiwddf = displaydf.copy()
            self.populatepiwdfilters(displaydf)
            self.displaypiwdtable(displaydf)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"PIWD Report failed: {str(e)}")
 
    def populatealertfilters(self, alertdf):
        if 'SCC' in alertdf.columns:
            self.alerts_scc_filter.additems(alertdf['SCC'].dropna().unique())
 
        if 'Alerts' in alertdf.columns:
            self.alerts_type_filter.additems(alertdf['Alerts'].dropna().unique())
 
        if 'Alerts' in alertdf.columns and 'Part' in alertdf.columns:
            identifiers = (alertdf['Alerts'].astype(str) + ' - ' + alertdf['Part'].astype(str)).dropna().unique()
            self.alerts_part_filter.additems(identifiers)

        if 'Region' in alertdf.columns:
            self.alerts_region_filter.additems(alertdf['Region'].dropna().unique())

        if 'Country' in alertdf.columns:
            self.alerts_country_filter.additems(alertdf['Country'].dropna().unique())

        if 'Program Supported' in alertdf.columns:
            self.alerts_program_filter.additems(alertdf['Program Supported'].dropna().unique())

    def applyalertfilters(self):
        if not hasattr(self, 'originalalertsdf'):
            return
 
        filtereddf = self.originalalertsdf.copy()
 
        selected_scc = self.alerts_scc_filter.getselecteditems()
        if selected_scc and 'SCC' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['SCC'].isin(selected_scc)]
 
        selected_alerts = self.alerts_type_filter.getselecteditems()
        if selected_alerts and 'Alerts' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['Alerts'].isin(selected_alerts)]
 
        selected_parts = self.alerts_part_filter.getselecteditems()
        if selected_parts and 'Alerts' in filtereddf.columns and 'Part' in filtereddf.columns:
            identifier_series = (filtereddf['Alerts'].astype(str) + ' - ' + filtereddf['Part'].astype(str))
            filtereddf = filtereddf[identifier_series.isin(selected_parts)]

        selected_regions = self.alerts_region_filter.getselecteditems()
        if selected_regions and 'Region' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['Region'].isin(selected_regions)]

        selected_countries = self.alerts_country_filter.getselecteditems()
        if selected_countries and 'Country' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['Country'].isin(selected_countries)]

        selected_programs = self.alerts_program_filter.getselecteditems()
        if selected_programs and 'Program Supported' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['Program Supported'].isin(selected_programs)]

        # Repopulate editable columns from the in-memory cache so any edits
        # the user made during this session are visible after a filter change,
        # even if the originalalertsdf sync in _onalertchanged missed a row.
        if self._alerts_cache and 'Part' in filtereddf.columns:
            for idx, row in filtereddf.iterrows():
                key = re.sub(r'\.0$', '', str(row['Part']).strip())
                if key in self._alerts_cache:
                    for col, val in self._alerts_cache[key].items():
                        if col in filtereddf.columns:
                            filtereddf.at[idx, col] = val

        self.displayalertstable(filtereddf)
 
    def clearalertfilters(self):
        if hasattr(self, 'alerts_scc_filter'):
            self.alerts_scc_filter.selectallitems()
        if hasattr(self, 'alerts_type_filter'):
            self.alerts_type_filter.selectallitems()
        if hasattr(self, 'alerts_part_filter'):
            self.alerts_part_filter.selectallitems()
        if hasattr(self, 'alerts_region_filter'):
            self.alerts_region_filter.selectallitems()
        if hasattr(self, 'alerts_country_filter'):
            self.alerts_country_filter.selectallitems()
        if hasattr(self, 'alerts_program_filter'):
            self.alerts_program_filter.selectallitems()
        if hasattr(self, 'originalalertsdf'):
            self.displayalertstable(self.originalalertsdf)
            
    def populatepiwdfilters(self, piwddf):
        if 'SCC Name' in piwddf.columns:
            self.piwd_scc_filter.additems(piwddf['SCC Name'].dropna().unique())
        
        if 'Part' in piwddf.columns:
            self.piwd_part_filter.additems(piwddf['Part'].dropna().unique())
            
    def applypiwdfilters(self):
        if not hasattr(self, 'originalpiwddf'):
            return
        
        filtereddf = self.originalpiwddf.copy()
        
        selected_scc = self.piwd_scc_filter.getselecteditems()
        if selected_scc and 'SCC Name' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['SCC Name'].isin(selected_scc)]
            
        selected_part = self.piwd_part_filter.getselecteditems()
        if selected_part and 'Part' in filtereddf.columns:
            filtereddf = filtereddf[filtereddf['Part'].isin(selected_part)]
            
        self.displaypiwdtable(filtereddf)
        
    def clearpiwdfilters(self):
        if hasattr(self, 'piwd_scc_filter'):
            self.piwd_scc_filter.selectallitems()
        if hasattr(self, 'piwd_part_filter'):
            self.piwd_part_filter.selectallitems()
 
    def displayalertstable(self, alertdf: pd.DataFrame):
        if alertdf.empty:
            return

        try:
            self.alertstable.itemChanged.disconnect()
        except Exception:
            pass

        cols = alertdf.columns.tolist()
        n_rows, n_cols = len(alertdf), len(cols)
        editable_indices = {i for i, c in enumerate(cols) if c in self._ALERT_EDITABLE_COLS}
        is_admin = self.userdata.get('username', '') in ADMINUSERS
        alerts_col_idx = cols.index('Alerts') if 'Alerts' in cols else -1
        part_col_idx = cols.index('Part') if 'Part' in cols else -1

        if self._alert_highlights is None:
            self._alert_highlights = self._loadalerthighlights()
        highlights = self._alert_highlights

        self.alertstable.setSortingEnabled(False)
        self.alertstable.setUpdatesEnabled(False)

        try:
            self.alertstable.setColumnCount(n_cols)
            self.alertstable.setHorizontalHeaderLabels(cols)
            self.alertstable.setRowCount(n_rows)

            data = alertdf.values
            int_col_indices = {i for i, c in enumerate(cols) if c in {'Inv', 'Yard', 'Req'}}

            for row in range(n_rows):
                part_str = str(data[row, part_col_idx]) if part_col_idx >= 0 else ''
                is_highlighted = part_str in highlights

                for col in range(n_cols):
                    value = data[row, col]
                    if col in int_col_indices and value is not None and not (isinstance(value, float) and value != value):
                        try:
                            display_value = str(int(float(value)))
                        except (ValueError, TypeError):
                            display_value = str(value)
                    else:
                        display_value = str(value) if value is not None and not (isinstance(value, float) and value != value) else ''

                    item = QTableWidgetItem(display_value)
                    if col in editable_indices:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                        item.setBackground(QColor(255, 255, 230))
                    elif col == alerts_col_idx:
                        if is_admin:
                            item.setFlags(item.flags() | Qt.ItemIsEditable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setBackground(QColor(255, 235, 59) if is_highlighted else QColor(230, 245, 255) if is_admin else QColor(255, 255, 255))
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                    self.alertstable.setItem(row, col, item)
 
        finally:
            self.alertstable.setUpdatesEnabled(True)
            self.alertstable.resizeColumnsToContents()
            self.alertstable.resizeRowsToContents()
            # Hide Region column — kept in data for filtering, not needed visually
            for c in range(self.alertstable.columnCount()):
                h = self.alertstable.horizontalHeaderItem(c)
                if h and h.text() == 'Region':
                    self.alertstable.setColumnHidden(c, True)
                    break
            self.alertstable.setSortingEnabled(True)
            self.alertstable.itemChanged.connect(self._onalertchanged)

    def exportalertstable(self):
        if not hasattr(self, 'originalalertsdf') or self.originalalertsdf.empty:
            QMessageBox.warning(self, "No Data", "Generate alerts breakdown first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Alerts Breakdown",
            f"Alerts_Breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx);;CSV files (*.csv)"
        )
        if not filename:
            return

        try:
            visible_cols = [c for c in range(self.alertstable.columnCount())
                            if not self.alertstable.isColumnHidden(c)]
            cols = [self.alertstable.horizontalHeaderItem(c).text() if self.alertstable.horizontalHeaderItem(c) else f'Col{c}'
                    for c in visible_cols]
            rows = []
            for r in range(self.alertstable.rowCount()):
                rows.append([self.alertstable.item(r, c).text() if self.alertstable.item(r, c) else ''
                              for c in visible_cols])
            df = pd.DataFrame(rows, columns=cols)

            if filename.endswith('.csv'):
                df.to_csv(filename, index=False)
            else:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                part_col_idx = cols.index('Part') if 'Part' in cols else -1
                highlights = self._alert_highlights or set()
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Alerts Breakdown')
                    ws = writer.sheets['Alerts Breakdown']
                    from openpyxl.styles import PatternFill
                    yellow_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
                    for row_idx, row_data in enumerate(rows):
                        part = row_data[part_col_idx] if part_col_idx >= 0 else ''
                        if part in highlights:
                            for col_idx in range(len(cols)):
                                ws.cell(row=row_idx + 2, column=col_idx + 1).fill = yellow_fill

            QMessageBox.information(self, "Export Complete", f"Exported to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))
                
    def displaypiwdtable(self, piwddf: pd.DataFrame):
        if piwddf.empty:
            return

        try:
            self.piwdtable.itemChanged.disconnect()
        except Exception:
            pass

        cols = piwddf.columns.tolist()
        n_rows, n_cols = len(piwddf), len(cols)
        editable_indices = {i for i, c in enumerate(cols) if c in self._PIWD_EDITABLE_COLS}

        self.piwdtable.setSortingEnabled(False)
        self.piwdtable.setUpdatesEnabled(False)
        
        try:
            self.piwdtable.setColumnCount(n_cols)
            self.piwdtable.setHorizontalHeaderLabels(cols)
            self.piwdtable.setRowCount(n_rows)
            
            data = piwddf.values
            
            int_col_indices = {i for i, c in enumerate(cols) if c in {'Inv', 'Yard', 'Req', 'Port'}}
            for row in range(n_rows):
                for col in range(n_cols):
                    value = data[row, col]
                    if col in int_col_indices and value is not None and not (isinstance(value, float) and value != value):
                        try:
                            display_value = str(int(float(value)))
                        except (ValueError, TypeError):
                            display_value = str(value)
                    else:
                        display_value = (str(value) if value is not None and not (isinstance(value, float) and value != value) else '')
                    item = QTableWidgetItem(display_value)
                    
                    if col in editable_indices:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                        item.setBackground(QColor(255, 255, 230))
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        
                    self.piwdtable.setItem(row, col, item)
                    
        finally:
            self.piwdtable.setUpdatesEnabled(True)
            self.piwdtable.resizeColumnsToContents()
            self.piwdtable.resizeRowsToContents()
            self.piwdtable.setSortingEnabled(True)
            self.piwdtable.itemChanged.connect(self._onpiwdchanged)

    def exportpiwdreport(self):
        if not hasattr(self, 'originalpiwddf') or self.originalpiwddf.empty:
            QMessageBox.warning(self, "No Data", "Generate PIWD report first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export PIWD Report",
            f"PIWD_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx);;CSV files (*.csv)"
        )
        if not filename:
            return
        try:
            if filename.endswith('.csv'):
                cols = [self.piwdtable.horizontalHeaderItem(c).text() if self.piwdtable.horizontalHeaderItem(c) else f'Col{c}' for c in range(self.piwdtable.columnCount())]
                rows = [[self.piwdtable.item(r, c).text() if self.piwdtable.item(r, c) else '' for c in range(self.piwdtable.columnCount())] for r in range(self.piwdtable.rowCount())]
                pd.DataFrame(rows, columns=cols).to_csv(filename, index=False)
            else:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                self._export_table_to_excel(self.piwdtable, filename, 'PIWD Report')
            QMessageBox.information(self, "Export Complete", f"Exported to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))
 
    def refreshcoveragedata(self):
        self.generatecoverageanalysis()
 
    def exportcoveragetable(self):
        if not hasattr(self, 'currentcoveragedf') or self.currentcoveragedf.empty:
            QMessageBox.warning(self, "No Data", "Generate analysis first.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Coverage Analysis",
            f"Coverage_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx);;CSV files (*.csv)"
        )
        if not filename:
            return
        try:
            if filename.endswith('.csv'):
                success = self.coverageengine.exporttocsv(self.currentcoveragedf, filename)
                if not success:
                    QMessageBox.critical(self, "Export Failed", "Could not export file.")
                    return
            else:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                self._export_table_to_excel(self.coveragetable, filename, 'Coverage Analysis')
            QMessageBox.information(self, "Export Complete", f"Exported to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    def resetcoveragetablesort(self):
        if not hasattr(self, 'currentcoveragedf') or self.currentcoveragedf.empty:
            return

        self.coveragetable.setSortingEnabled(False)
        self.displaycoveragetable(self.currentcoveragedf)

        header = self.coveragetable.horizontalHeader()
        try:
            header.setSortIndicator(-1, Qt.AscendingOrder)
        except Exception:
            pass
 
    def closeEvent(self, event):
        event.accept()
